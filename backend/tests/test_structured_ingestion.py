from __future__ import annotations

import tempfile
import unittest
from dataclasses import replace
from decimal import Decimal
from pathlib import Path
from threading import Event, Thread

from app.structured_ingestion import (
    ArrowParquetSink,
    SpreadsheetPublisher,
    StructuredIngestionError,
)
from app.structured_models import StructuredColumnType, StructuredPublicationResult
from tests.support.structured_fakes import (
    RecordingParquetSink,
    sample_confirmed_schema,
    write_csv,
    write_formula_xlsx,
    write_xlsx,
)


class RecordingPublicationGateway:
    def __init__(self) -> None:
        self.prepared: list[tuple[object, str, str]] = []
        self.inserted_batch_rows: list[int] = []
        self.validations: list[dict[str, object]] = []

    def prepare_publication(self, schema, publication_id: str, content_hash: str):
        self.prepared.append((schema, publication_id, content_hash))
        return type(
            "Target",
            (),
            {
                "staging_table": "structured_ds_sales_v1_staging",
                "physical_table_name": "structured_ds_sales_v1",
                "content_hash": content_hash,
            },
        )()

    def insert_batch(self, target, batch) -> None:
        self.inserted_batch_rows.append(batch.num_rows)
        self.assert_content_hash_column(target.content_hash, batch)

    @staticmethod
    def assert_content_hash_column(expected: str, batch) -> None:
        values = batch.column(batch.schema.get_field_index("_content_hash")).to_pylist()
        if values != [expected] * batch.num_rows:
            raise AssertionError(values)

    def validate_and_promote(self, target, **statistics) -> str:
        self.validations.append(dict(statistics))
        return target.physical_table_name


class FailingPublicationGateway(RecordingPublicationGateway):
    def __init__(self) -> None:
        super().__init__()
        self.discarded: list[object] = []

    def validate_and_promote(self, target, **statistics) -> str:
        self.validations.append(dict(statistics))
        raise RuntimeError("validation failed")

    def discard_publication(self, target) -> None:
        self.discarded.append(target)


class CoordinatedAttemptSink(ArrowParquetSink):
    def __init__(self, root: Path, blocked_attempt_directory: str) -> None:
        super().__init__(root)
        self.blocked_attempt_directory = blocked_attempt_directory
        self.blocked_attempt_written = Event()
        self.release_blocked_attempt = Event()

    def write_batch(self, batch, output_path: Path) -> None:
        super().write_batch(batch, output_path)
        if output_path.parent.name == self.blocked_attempt_directory:
            self.blocked_attempt_written.set()
            if not self.release_blocked_attempt.wait(timeout=5):
                raise TimeoutError("blocked publication attempt was not released")


class StructuredIngestionTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir_context = tempfile.TemporaryDirectory()
        self.temp_dir = Path(self.temp_dir_context.name)

    def tearDown(self) -> None:
        self.temp_dir_context.cleanup()

    def test_ingestion_writes_bounded_batches_and_counts_rows(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=5)
        sink = RecordingParquetSink(self.temp_dir / "parquet")
        gateway = RecordingPublicationGateway()

        result = SpreadsheetPublisher(
            sink=sink,
            clickhouse=gateway,
            batch_rows=2,
        ).publish(
            path=confirmed.path,
            schema=confirmed.schema,
            publication_id="pub-1",
        )

        self.assertIsInstance(result, StructuredPublicationResult)
        self.assertEqual(result.row_count, 5)
        self.assertEqual(result.column_count, 3)
        self.assertEqual(result.null_counts, {"order_amount": 0, "region": 0, "order_date": 0})
        self.assertLessEqual(max(sink.batch_rows), 2)
        self.assertLessEqual(max(gateway.inserted_batch_rows), 2)
        self.assertEqual(sum(gateway.inserted_batch_rows), 5)
        self.assertEqual(len(result.content_hash), 64)
        self.assertTrue(
            all(
                path.parts[-5:-1] == ("kb-sales", "ds-sales", "1", "pub-1")
                for path in sink.output_paths
            )
        )
        self.assertEqual(gateway.validations[0]["row_count"], 5)
        self.assertEqual(gateway.validations[0]["content_hash"], result.content_hash)

    def test_ingestion_flushes_variable_width_rows_at_the_byte_limit(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        path = write_csv(
            self.temp_dir / "wide.csv",
            [
                [column.original_name for column in confirmed.schema.columns],
                *[[index, "x" * 400, "2026-01-01"] for index in range(1, 4)],
            ],
        )
        schema = replace(confirmed.schema, worksheet_name=path.stem)
        sink = RecordingParquetSink(self.temp_dir / "parquet-wide")

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
            batch_rows=50_000,
            batch_bytes=512,
        ).publish(path, schema, "pub-wide")

        self.assertEqual(sink.batch_rows, [1, 1, 1])

    def test_single_row_larger_than_the_byte_limit_is_rejected(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        path = write_csv(
            self.temp_dir / "oversized-row.csv",
            [
                [column.original_name for column in confirmed.schema.columns],
                ["1", "secret-" + "x" * 2_000, "2026-01-01"],
            ],
        )
        schema = replace(confirmed.schema, worksheet_name=path.stem)

        with self.assertRaises(StructuredIngestionError) as raised:
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet-oversized"),
                clickhouse=RecordingPublicationGateway(),
                batch_bytes=512,
            ).publish(path, schema, "pub-oversized")

        self.assertEqual(raised.exception.code, "row_size_limit_exceeded")
        self.assertNotIn("secret-", str(raised.exception))

    def test_parquet_rows_include_stable_provenance_and_row_numbers(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=3)
        sink = RecordingParquetSink(self.temp_dir / "parquet")

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
            batch_rows=2,
        ).publish(confirmed.path, confirmed.schema, "pub-2")

        rows = [row for batch in sink.batches for row in batch.to_pylist()]
        self.assertEqual([row["_row_number"] for row in rows], [2, 3, 4])
        self.assertTrue(all(row["_source_id"] == "kb-sales" for row in rows))
        self.assertTrue(all(row["_dataset_id"] == "ds-sales" for row in rows))
        self.assertTrue(all(row["_schema_version"] == 1 for row in rows))
        self.assertTrue(all(row["_worksheet"] == confirmed.schema.worksheet_name for row in rows))

    def test_parquet_path_components_cannot_escape_the_configured_root(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        schema = replace(confirmed.schema, source_id="..")
        root = self.temp_dir / "parquet"
        sink = RecordingParquetSink(root)

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
        ).publish(confirmed.path, schema, "pub-safe-path")

        self.assertTrue(sink.output_paths[0].resolve().is_relative_to(root.resolve()))

    def test_decimal_values_remain_exact_in_arrow_and_parquet_batches(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=2)
        sink = RecordingParquetSink(self.temp_dir / "parquet")

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
            batch_rows=2,
        ).publish(confirmed.path, confirmed.schema, "pub-decimal")

        values = [row["order_amount"] for batch in sink.batches for row in batch.to_pylist()]
        self.assertTrue(all(isinstance(value, Decimal) for value in values))
        self.assertEqual(values, [Decimal("10.000000000"), Decimal("20.000000000")])

    def test_decimal_zero_policy_writes_a_decimal_zero(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        zero_amount = replace(confirmed.schema.columns[0], null_policy="zero")
        schema = replace(
            confirmed.schema,
            columns=(zero_amount, *confirmed.schema.columns[1:]),
        )
        path = write_xlsx(
            self.temp_dir / "zero.xlsx",
            schema.worksheet_name,
            [
                [column.original_name for column in schema.columns],
                [None, "east", "2026-01-01"],
            ],
        )
        sink = RecordingParquetSink(self.temp_dir / "parquet")

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
        ).publish(path, schema, "pub-zero")

        amount = sink.batches[0].to_pylist()[0]["order_amount"]
        self.assertEqual(amount, Decimal("0.000000000"))
        self.assertIsInstance(amount, Decimal)

    def test_decimal_38_9_accepts_29_integer_digits_and_rejects_30(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        for integer_digits, accepted in ((29, True), (30, False)):
            with self.subTest(integer_digits=integer_digits):
                path = write_csv(
                    self.temp_dir / f"decimal-{integer_digits}.csv",
                    [
                        ["order_amount", "region", "order_date"],
                        [f"{'9' * integer_digits}.123456789", "east", "2026-01-01"],
                    ],
                )
                schema = replace(confirmed.schema, worksheet_name=path.stem)
                publisher = SpreadsheetPublisher(
                    sink=RecordingParquetSink(self.temp_dir / f"parquet-{integer_digits}"),
                    clickhouse=RecordingPublicationGateway(),
                )
                if accepted:
                    result = publisher.publish(path, schema, f"pub-{integer_digits}")
                    self.assertEqual(result.row_count, 1)
                else:
                    with self.assertRaises(StructuredIngestionError) as raised:
                        publisher.publish(path, schema, f"pub-{integer_digits}")
                        self.assertEqual(raised.exception.code, "type_conversion_failed")

    def test_timezone_aware_datetime_is_rejected_before_arrow_and_digest(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        datetime_column = replace(
            confirmed.schema.columns[2],
            data_type=StructuredColumnType.DATETIME,
        )
        schema = replace(
            confirmed.schema,
            worksheet_name="aware",
            columns=(*confirmed.schema.columns[:2], datetime_column),
        )
        path = write_csv(
            self.temp_dir / "aware.csv",
            [
                [column.original_name for column in schema.columns],
                ["10", "east", "2026-01-01T10:00:00+08:00"],
            ],
        )

        with self.assertRaises(StructuredIngestionError) as raised:
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet-aware"),
                clickhouse=RecordingPublicationGateway(),
            ).publish(path, schema, "pub-aware")

        self.assertEqual(raised.exception.code, "type_conversion_failed")
        self.assertEqual(raised.exception.physical_column, datetime_column.physical_name)

    def test_numeric_ranges_include_an_all_null_numeric_column(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        path = write_xlsx(
            self.temp_dir / "nulls.xlsx",
            confirmed.schema.worksheet_name,
            [
                [column.original_name for column in confirmed.schema.columns],
                [None, "east", "2026-01-01"],
            ],
        )
        gateway = RecordingPublicationGateway()

        SpreadsheetPublisher(
            sink=RecordingParquetSink(self.temp_dir / "parquet"),
            clickhouse=gateway,
        ).publish(path, confirmed.schema, "pub-nulls")

        self.assertEqual(
            gateway.validations[0]["numeric_ranges"],
            {"order_amount": (None, None)},
        )

    def test_maps_confirmed_columns_by_physical_header_not_submission_order(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        reordered = replace(
            confirmed.schema,
            columns=tuple(reversed(confirmed.schema.columns)),
        )
        sink = RecordingParquetSink(self.temp_dir / "parquet")

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
        ).publish(confirmed.path, reordered, "pub-reordered")

        row = sink.batches[0].to_pylist()[0]
        self.assertEqual(row["order_amount"], Decimal("10.000000000"))
        self.assertIsInstance(row["region"], str)
        self.assertTrue(row["region"])
        self.assertEqual(str(row["order_date"]), "2026-01-01")

    def test_rejects_source_header_drift_before_clickhouse_preparation(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        path = write_xlsx(
            self.temp_dir / "drift.xlsx",
            confirmed.schema.worksheet_name,
            [
                ["different_amount", "region", "order_date"],
                [10, "east", "2026-01-01"],
            ],
        )
        gateway = RecordingPublicationGateway()

        with self.assertRaises(StructuredIngestionError) as raised:
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet"),
                clickhouse=gateway,
            ).publish(path, confirmed.schema, "pub-drift")

        self.assertEqual(raised.exception.code, "schema_drift")
        self.assertEqual(gateway.prepared, [])

    def test_streams_a_confirmed_gb18030_csv_without_decoding_the_whole_file(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        path = write_csv(
            self.temp_dir / "sales.csv",
            [
                ["order_amount", "region", "order_date"],
                ["10.5", "华东", "2026-01-01"],
            ],
            encoding="gb18030",
        )
        schema = replace(confirmed.schema, worksheet_name=path.stem)
        sink = RecordingParquetSink(self.temp_dir / "parquet")

        result = SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
            batch_rows=1,
        ).publish(path, schema, "pub-csv")

        self.assertEqual(result.row_count, 1)
        row = sink.batches[0].to_pylist()[0]
        self.assertEqual(row["region"], "华东")

    def test_real_parquet_sink_round_trips_bounded_parts(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=3)
        sink = ArrowParquetSink(self.temp_dir / "parquet")

        result = SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
            batch_rows=2,
        ).publish(confirmed.path, confirmed.schema, "pub-parquet")

        parts = sorted((self.temp_dir / "parquet").rglob("*.parquet"))
        self.assertEqual(result.row_count, 3)
        self.assertEqual(len(parts), 2)
        round_tripped = [row for batch in sink.iter_batches(parts) for row in batch.to_pylist()]
        self.assertEqual([row["_row_number"] for row in round_tripped], [2, 3, 4])

    def test_conversion_failure_reports_row_column_and_redacted_sample(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=2)
        from openpyxl import load_workbook

        workbook = load_workbook(confirmed.path)
        worksheet = workbook[confirmed.schema.worksheet_name]
        worksheet.cell(row=3, column=1, value="not-a-number-secret")
        workbook.save(confirmed.path)
        workbook.close()

        with self.assertRaises(StructuredIngestionError) as raised:
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet"),
                clickhouse=RecordingPublicationGateway(),
                batch_rows=2,
            ).publish(confirmed.path, confirmed.schema, "pub-bad")

        error = raised.exception
        self.assertEqual(error.code, "type_conversion_failed")
        self.assertEqual(error.row_number, 3)
        self.assertEqual(error.physical_column, "order_amount")
        self.assertNotEqual(error.sample, "not-a-number-secret")
        self.assertIn("[redacted", error.sample)
        self.assertNotIn("not-a-number-secret", str(error))
        self.assertIn("[redacted", str(error))

    def test_formula_without_cached_value_blocks_aggregate_publication(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        formula_path = write_formula_xlsx(
            self.temp_dir / "formula.xlsx",
            header=confirmed.schema.columns[0].original_name,
            formula="=SUM(1,2)",
        )
        formula_schema = replace(
            confirmed.schema,
            worksheet_name="Sheet1",
            columns=(confirmed.schema.columns[0],),
        )

        with self.assertRaises(StructuredIngestionError) as raised:
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet"),
                clickhouse=RecordingPublicationGateway(),
            ).publish(formula_path, formula_schema, "pub-formula")

        self.assertEqual(raised.exception.code, "formula_cache_missing")
        self.assertIn("recalculate", str(raised.exception).lower())
        self.assertIn("save", str(raised.exception).lower())

    def test_nonaggregate_formula_without_cache_still_applies_null_policy(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        formula_path = write_formula_xlsx(
            self.temp_dir / "formula-null-policy.xlsx",
            header=confirmed.schema.columns[0].original_name,
            formula="=SUM(1,2)",
        )
        base_column = replace(confirmed.schema.columns[0], allow_aggregate=False)

        zero_schema = replace(
            confirmed.schema,
            worksheet_name="Sheet1",
            columns=(replace(base_column, null_policy="zero"),),
        )
        sink = RecordingParquetSink(self.temp_dir / "parquet-zero")
        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
        ).publish(formula_path, zero_schema, "pub-formula-zero")
        self.assertEqual(
            sink.batches[0].to_pylist()[0]["order_amount"],
            Decimal("0.000000000"),
        )

        reject_schema = replace(
            confirmed.schema,
            worksheet_name="Sheet1",
            columns=(replace(base_column, null_policy="reject"),),
        )
        with self.assertRaises(StructuredIngestionError) as raised:
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet-reject"),
                clickhouse=RecordingPublicationGateway(),
            ).publish(formula_path, reject_schema, "pub-formula-reject")
        self.assertEqual(raised.exception.code, "null_value_rejected")

    def test_clickhouse_failure_discards_staging_for_same_publication_retry(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        gateway = FailingPublicationGateway()

        with self.assertRaisesRegex(RuntimeError, "validation failed"):
            SpreadsheetPublisher(
                sink=RecordingParquetSink(self.temp_dir / "parquet"),
                clickhouse=gateway,
            ).publish(confirmed.path, confirmed.schema, "pub-retry")

        self.assertEqual(len(gateway.discarded), 1)
        self.assertEqual(gateway.discarded[0].staging_table, "structured_ds_sales_v1_staging")

    def test_same_publication_retry_removes_stale_parquet_parts(self) -> None:
        first = sample_confirmed_schema(self.temp_dir, row_count=5)
        sink = ArrowParquetSink(self.temp_dir / "parquet")
        with self.assertRaisesRegex(RuntimeError, "validation failed"):
            SpreadsheetPublisher(
                sink=sink,
                clickhouse=FailingPublicationGateway(),
                batch_rows=2,
            ).publish(first.path, first.schema, "pub-retry")

        second = sample_confirmed_schema(self.temp_dir, row_count=1)
        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
            batch_rows=2,
        ).publish(second.path, second.schema, "pub-retry")

        publication_parts = list(
            (self.temp_dir / "parquet" / "kb-sales" / "ds-sales" / "1" / "pub-retry").glob(
                "*.parquet"
            )
        )
        self.assertEqual([path.name for path in publication_parts], ["part-00000.parquet"])

    def test_stale_attempt_cleanup_does_not_delete_takeover_attempt_parts(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        parquet_root = self.temp_dir / "parquet"
        sink = CoordinatedAttemptSink(parquet_root, "attempt-old-owner")
        lease_lost = Event()
        old_errors: list[Exception] = []

        def old_lease_guard() -> None:
            if lease_lost.is_set():
                raise RuntimeError("old lease lost")

        def publish_old_attempt() -> None:
            try:
                SpreadsheetPublisher(
                    sink=sink,
                    clickhouse=RecordingPublicationGateway(),
                ).publish(
                    confirmed.path,
                    confirmed.schema,
                    "pub-overlap",
                    lease_guard=old_lease_guard,
                    staging_token="old-owner",
                )
            except Exception as error:
                old_errors.append(error)

        old_thread = Thread(target=publish_old_attempt)
        old_thread.start()
        self.assertTrue(
            sink.blocked_attempt_written.wait(timeout=1),
            f"old attempt failed before its blocked write: {old_errors}",
        )

        SpreadsheetPublisher(
            sink=sink,
            clickhouse=RecordingPublicationGateway(),
        ).publish(
            confirmed.path,
            confirmed.schema,
            "pub-overlap",
            staging_token="new-owner",
        )
        new_part = (
            parquet_root
            / "kb-sales"
            / "ds-sales"
            / "1"
            / "pub-overlap"
            / "attempt-new-owner"
            / "part-00000.parquet"
        )
        self.assertTrue(new_part.exists())
        old_attempt = (
            parquet_root / "kb-sales" / "ds-sales" / "1" / "pub-overlap" / "attempt-old-owner"
        )
        self.assertFalse(old_attempt.exists())

        lease_lost.set()
        sink.release_blocked_attempt.set()
        old_thread.join(timeout=5)

        self.assertFalse(old_thread.is_alive())
        self.assertEqual([str(error) for error in old_errors], ["old lease lost"])
        self.assertTrue(new_part.exists())
        self.assertFalse(old_attempt.exists())

    def test_attempt_start_removes_only_stale_attempt_siblings(self) -> None:
        confirmed = sample_confirmed_schema(self.temp_dir, row_count=1)
        publication_root = self.temp_dir / "parquet" / "kb-sales" / "ds-sales" / "1" / "pub-gc"
        stale = publication_root / "attempt-stale-owner"
        legacy = publication_root / "legacy-output"
        stale.mkdir(parents=True)
        legacy.mkdir()
        (stale / "part-00000.parquet").write_bytes(b"stale")
        (legacy / "keep.txt").write_text("keep", encoding="utf-8")

        SpreadsheetPublisher(
            sink=ArrowParquetSink(self.temp_dir / "parquet"),
            clickhouse=RecordingPublicationGateway(),
        ).publish(
            confirmed.path,
            confirmed.schema,
            "pub-gc",
            staging_token="current-owner",
        )

        self.assertFalse(stale.exists())
        self.assertTrue(legacy.exists())
        self.assertTrue(
            (publication_root / "attempt-current-owner" / "part-00000.parquet").exists()
        )


if __name__ == "__main__":
    unittest.main()
