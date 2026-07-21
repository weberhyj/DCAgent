from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

import app.spreadsheet_schema as schema_module
from app.spreadsheet_schema import _infer_dataset, infer_spreadsheet_schema
from app.structured_models import StructuredColumnType
from tests.support.structured_fakes import write_csv, write_formula_xlsx, write_xlsx


class SpreadsheetSchemaInferenceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir_context = tempfile.TemporaryDirectory()
        self.temp_dir = Path(self.temp_dir_context.name)

    def tearDown(self) -> None:
        self.temp_dir_context.cleanup()

    def test_infers_numeric_column_and_aliases_from_xlsx(self) -> None:
        path = write_xlsx(
            self.temp_dir / "sales.xlsx",
            "明细",
            [
                ["订单金额", "地区", "日期"],
                ["10.5", "华东", "2026-01-01"],
                ["20", "华南", "2026-01-02"],
            ],
        )

        preview = infer_spreadsheet_schema(path, source_id="kb-sales")

        self.assertEqual(len(preview.datasets), 1)
        dataset = preview.datasets[0]
        self.assertEqual(dataset.worksheet_name, "明细")
        self.assertEqual(dataset.sampled_rows, 2)
        self.assertIs(dataset.columns[0].data_type, StructuredColumnType.DECIMAL)
        self.assertIn("订单金额", dataset.columns[0].aliases)
        self.assertEqual(dataset.columns[0].examples, ("10.5", "20"))
        self.assertIs(dataset.columns[2].data_type, StructuredColumnType.DATE)

    def test_mixed_numeric_values_upgrade_to_string_and_report_rows(self) -> None:
        path = write_xlsx(self.temp_dir / "mixed.xlsx", "Sheet1", [["金额"], ["10"], ["未知"]])

        preview = infer_spreadsheet_schema(path, source_id="kb-mixed")

        self.assertIs(preview.datasets[0].columns[0].data_type, StructuredColumnType.STRING)
        diagnostic = next(item for item in preview.diagnostics if item.code == "mixed_type")
        self.assertEqual(diagnostic.worksheet_name, "Sheet1")
        self.assertEqual(diagnostic.column_name, "金额")
        self.assertEqual(diagnostic.row_number, 3)

    def test_duplicate_headers_are_stable_and_require_confirmation(self) -> None:
        path = write_csv(self.temp_dir / "duplicate.csv", [["amount", "amount"], ["10", "20"]])

        preview = infer_spreadsheet_schema(path, source_id="kb-duplicate")

        self.assertEqual(
            [column.physical_name for column in preview.datasets[0].columns],
            ["amount", "amount_2"],
        )
        diagnostic = next(item for item in preview.diagnostics if item.code == "duplicate_header")
        self.assertEqual(diagnostic.column_name, "amount")
        self.assertEqual(diagnostic.row_number, 1)

    def test_formula_without_cached_value_is_reported(self) -> None:
        path = write_formula_xlsx(
            self.temp_dir / "formula.xlsx", header="合计", formula="=SUM(1,2)"
        )

        preview = infer_spreadsheet_schema(path, source_id="kb-formula")

        self.assertEqual(len(preview.datasets), 1)
        matching = [item for item in preview.diagnostics if item.code == "formula_cache_missing"]
        self.assertEqual(len(matching), 1)
        diagnostic = matching[0]
        self.assertEqual(diagnostic.worksheet_name, "Sheet1")
        self.assertEqual(diagnostic.column_name, "合计")
        self.assertEqual(diagnostic.row_number, 2)

    def test_formula_without_cached_header_value_is_reported(self) -> None:
        path = self.temp_dir / "formula-header.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["=A1"])
        sheet.append(["value"])
        workbook.save(path)
        workbook.close()

        preview = infer_spreadsheet_schema(path, source_id="kb-formula-header")

        matching = [item for item in preview.diagnostics if item.code == "formula_cache_missing"]
        self.assertEqual(len(matching), 1)
        diagnostic = matching[0]
        self.assertEqual(diagnostic.worksheet_name, "Sheet1")
        self.assertEqual(diagnostic.column_name, "column_1")
        self.assertEqual(diagnostic.row_number, 1)

    def test_sampling_cap_does_not_consume_an_extra_data_row(self) -> None:
        consumed: list[int] = []

        def rows():
            consumed.append(1)
            yield (1, ("value",), ())
            for row_number in range(2, 10_002):
                consumed.append(row_number)
                yield (row_number, (str(row_number),), ())
            consumed.append(10_002)
            yield (10_002, ("must-not-be-read",), ())

        diagnostics = []
        dataset = _infer_dataset("kb-cap", "Sheet1", rows(), diagnostics)

        self.assertIsNotNone(dataset)
        assert dataset is not None
        self.assertEqual(dataset.sampled_rows, 10_000)
        self.assertEqual(len(consumed), 10_001)
        self.assertNotIn(10_002, consumed)

    def test_excel_number_format_distinguishes_date_from_midnight_datetime(self) -> None:
        path = self.temp_dir / "midnight.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Types"
        sheet.append(["day", "timestamp"])
        sheet.append([date(2026, 1, 1), datetime(2026, 1, 1, 0, 0, 0)])
        sheet["A2"].number_format = "yyyy-mm-dd"
        sheet["B2"].number_format = "yyyy-mm-dd h:mm:ss"
        workbook.save(path)
        workbook.close()

        preview = infer_spreadsheet_schema(path, source_id="kb-midnight")

        self.assertEqual(
            [column.data_type for column in preview.datasets[0].columns],
            [StructuredColumnType.DATE, StructuredColumnType.DATETIME],
        )

    def test_empty_sheet_is_blocking(self) -> None:
        path = write_xlsx(self.temp_dir / "empty.xlsx", "Sheet1", [])

        preview = infer_spreadsheet_schema(path, source_id="kb-empty")

        self.assertEqual(preview.datasets, ())
        self.assertTrue(any(item.code == "empty_sheet" for item in preview.diagnostics))

    def test_unsupported_csv_encoding_is_blocking(self) -> None:
        path = self.temp_dir / "unsupported.csv"
        path.write_bytes(b"\xff\xfe\x00\x00")

        preview = infer_spreadsheet_schema(path, source_id="kb-encoding")

        self.assertEqual(preview.datasets, ())
        self.assertTrue(any(item.code == "unsupported_encoding" for item in preview.diagnostics))

    def test_streams_supported_csv_encodings(self) -> None:
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            with self.subTest(encoding=encoding):
                path = write_csv(
                    self.temp_dir / f"sales-{encoding}.csv",
                    [["金额", "日期"], ["10", "2026-01-01"]],
                    encoding=encoding,
                )

                preview = infer_spreadsheet_schema(path, source_id=f"kb-{encoding}")

                self.assertEqual(len(preview.datasets), 1)
                self.assertIs(
                    preview.datasets[0].columns[0].data_type,
                    StructuredColumnType.INTEGER,
                )
                self.assertIs(
                    preview.datasets[0].columns[1].data_type,
                    StructuredColumnType.DATE,
                )

    def test_returns_each_non_empty_worksheet_with_stable_ids_and_hashes(self) -> None:
        path = write_xlsx(
            self.temp_dir / "multi.xlsx",
            {
                "North": [["amount"], ["10"]],
                "Empty": [],
                "South": [["amount"], ["20"]],
            },
        )

        first = infer_spreadsheet_schema(path, source_id="kb-multi")
        second = infer_spreadsheet_schema(path, source_id="kb-multi")

        self.assertEqual([dataset.worksheet_name for dataset in first.datasets], ["North", "South"])
        self.assertEqual(
            [dataset.dataset_id for dataset in first.datasets],
            [dataset.dataset_id for dataset in second.datasets],
        )
        self.assertEqual(
            [dataset.schema_hash for dataset in first.datasets],
            [dataset.schema_hash for dataset in second.datasets],
        )
        self.assertEqual(first.datasets[0].schema_hash, first.datasets[1].schema_hash)
        self.assertTrue(
            any(
                item.code == "empty_sheet" and item.worksheet_name == "Empty"
                for item in first.diagnostics
            )
        )

    def test_generates_blank_and_non_ascii_physical_headers(self) -> None:
        path = write_xlsx(
            self.temp_dir / "headers.xlsx",
            "Sheet1",
            [[None, "123 Total", "订单金额"], ["x", 1, 2]],
        )

        preview = infer_spreadsheet_schema(path, source_id="kb-headers")

        columns = preview.datasets[0].columns
        self.assertEqual(
            [column.physical_name for column in columns],
            ["column_1", "col_123_total", "column_3"],
        )
        self.assertEqual(columns[0].original_name, "")
        self.assertEqual(columns[0].display_name, "column_1")
        self.assertEqual(columns[0].aliases, ())
        self.assertEqual(columns[2].original_name, "订单金额")
        self.assertIn("订单金额", columns[2].aliases)
        self.assertTrue(any(item.code == "missing_header" for item in preview.diagnostics))

    def test_caps_examples_and_sampled_rows(self) -> None:
        rows = [["value"], *[[str(index)] for index in range(10_005)]]
        path = write_csv(self.temp_dir / "large.csv", rows)

        preview = infer_spreadsheet_schema(path, source_id="kb-large")

        dataset = preview.datasets[0]
        self.assertEqual(dataset.sampled_rows, 10_000)
        self.assertEqual(dataset.columns[0].sampled_rows, 10_000)
        self.assertEqual(dataset.columns[0].examples, ("0", "1", "2", "3", "4"))

    def test_caps_wide_rows_and_reports_column_truncation(self) -> None:
        max_columns = 256
        width = max_columns + 10
        path = write_csv(
            self.temp_dir / "wide.csv",
            [[f"column_{index}" for index in range(width)], list(range(width))],
        )

        preview = infer_spreadsheet_schema(path, source_id="kb-wide")

        self.assertEqual(len(preview.datasets[0].columns), max_columns)
        self.assertTrue(any(item.code == "column_limit_exceeded" for item in preview.diagnostics))

    def test_caps_diagnostics_and_reports_truncation(self) -> None:
        max_columns = 256
        max_diagnostics = 128
        header = tuple("amount" for _ in range(max_columns))
        diagnostics = []

        dataset = _infer_dataset("kb-diagnostics", "Sheet1", [(1, header, ())], diagnostics)

        self.assertIsNotNone(dataset)
        self.assertEqual(len(diagnostics), max_diagnostics)
        self.assertEqual(diagnostics[-1].code, "diagnostics_truncated")

    def test_csv_encoding_probe_reads_only_a_bounded_prefix(self) -> None:
        payload = b"header\n" + (b"value\n" * 100_000)

        class CountingPath:
            def __init__(self) -> None:
                self.bytes_read = 0

            def open(self, mode: str):
                self.assert_binary_mode(mode)

                owner = self

                class CountingStream(BytesIO):
                    def read(self, size: int = -1):
                        chunk = super().read(size)
                        owner.bytes_read += len(chunk)
                        return chunk

                return CountingStream(payload)

            @staticmethod
            def assert_binary_mode(mode: str) -> None:
                if mode != "rb":
                    raise AssertionError(mode)

        path = CountingPath()

        encoding = schema_module._detect_csv_encoding(path)

        self.assertEqual(encoding, "utf-8-sig")
        self.assertLessEqual(path.bytes_read, 64 * 1024)

    def test_rejects_nul_prefixed_csv_as_unsupported_encoding(self) -> None:
        path = self.temp_dir / "utf16.csv"
        path.write_bytes("amount,region\n10,east\n".encode("utf-16-le"))

        preview = infer_spreadsheet_schema(path, source_id="kb-utf16")

        self.assertEqual(preview.datasets, ())
        self.assertTrue(any(item.code == "unsupported_encoding" for item in preview.diagnostics))

    def test_sheet_read_error_preserves_an_already_valid_sheet(self) -> None:
        class Cell:
            def __init__(self, value):
                self.value = value
                self.number_format = "General"

        class Sheet:
            def __init__(self, title, rows=(), error=None):
                self.title = title
                self.rows = rows
                self.error = error

            def iter_rows(self):
                if self.error is not None:
                    raise self.error
                return iter(tuple(tuple(Cell(value) for value in row) for row in self.rows))

        class WorkbookStub:
            def __init__(self, sheets):
                self.worksheets = sheets

            def close(self):
                return None

        cached = WorkbookStub(
            [
                Sheet("Good", (("amount",), (10,))),
                Sheet("Bad", error=OSError("broken sheet stream")),
            ]
        )
        formulas = WorkbookStub([Sheet("Good", (("amount",), (10,))), Sheet("Bad", (("amount",),))])

        with patch.object(schema_module, "load_workbook", side_effect=(cached, formulas)):
            preview = infer_spreadsheet_schema(self.temp_dir / "book.xlsx", "kb-book")

        self.assertEqual([dataset.worksheet_name for dataset in preview.datasets], ["Good"])
        self.assertTrue(
            any(
                item.code == "sheet_read_error" and item.worksheet_name == "Bad"
                for item in preview.diagnostics
            )
        )

    def test_infers_dates_datetimes_booleans_and_null_counts(self) -> None:
        path = write_xlsx(
            self.temp_dir / "types.xlsx",
            "Types",
            [
                ["day", "timestamp", "enabled", "optional"],
                [date(2026, 1, 1), datetime(2026, 1, 1, 9, 30), True, None],
                [date(2026, 1, 2), datetime(2026, 1, 2, 10, 45), False, "present"],
            ],
        )

        preview = infer_spreadsheet_schema(path, source_id="kb-types")

        columns = preview.datasets[0].columns
        self.assertEqual(
            [column.data_type for column in columns],
            [
                StructuredColumnType.DATE,
                StructuredColumnType.DATETIME,
                StructuredColumnType.BOOLEAN,
                StructuredColumnType.STRING,
            ],
        )
        self.assertEqual([column.null_count for column in columns], [0, 0, 0, 1])
        self.assertEqual(len(preview.datasets[0].schema_hash), 64)
        self.assertFalse(any(item.code == "mixed_type" for item in preview.diagnostics))

    def test_empty_csv_is_blocking(self) -> None:
        path = self.temp_dir / "empty.csv"
        path.write_text("", encoding="utf-8")

        preview = infer_spreadsheet_schema(path, source_id="kb-empty-csv")

        self.assertEqual(preview.datasets, ())
        self.assertTrue(any(item.code == "empty_sheet" for item in preview.diagnostics))

    def test_rejects_unsupported_spreadsheet_suffix(self) -> None:
        path = self.temp_dir / "sales.xls"
        path.write_bytes(b"legacy")

        with self.assertRaisesRegex(ValueError, "Unsupported spreadsheet format"):
            infer_spreadsheet_schema(path, source_id="kb-legacy")


if __name__ == "__main__":
    unittest.main()
