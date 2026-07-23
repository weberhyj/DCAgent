from __future__ import annotations

import csv
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.structured_models import (
    StructuredCatalog,
    StructuredColumnSchema,
    StructuredColumnType,
    StructuredDatasetCatalog,
    StructuredDatasetSchema,
    StructuredPublication,
)

Rows = Sequence[Sequence[Any]]


def write_xlsx(
    path: Path,
    worksheet_name: str | Mapping[str, Rows],
    rows: Rows | None = None,
) -> Path:
    """Write exactly the supplied worksheet names and row values to an XLSX file."""
    workbook = Workbook()
    default_sheet = workbook.active
    if isinstance(worksheet_name, Mapping):
        worksheets = worksheet_name
    else:
        worksheets = {worksheet_name: rows or ()}

    for index, (name, worksheet_rows) in enumerate(worksheets.items()):
        sheet = default_sheet if index == 0 else workbook.create_sheet()
        sheet.title = name
        for row in worksheet_rows:
            sheet.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def write_formula_xlsx(path: Path, header: str, formula: str) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append([header])
    sheet.append([formula])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def write_csv(path: Path, rows: Rows, *, encoding: str = "utf-8") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding=encoding, newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)
    return path


@dataclass(frozen=True, slots=True)
class ConfirmedSpreadsheetFixture:
    path: Path
    schema: StructuredDatasetSchema


def sample_columns() -> tuple[StructuredColumnSchema, ...]:
    return (
        StructuredColumnSchema(
            physical_name="order_amount",
            original_name="订单金额",
            display_name="订单金额",
            data_type=StructuredColumnType.DECIMAL,
            aliases=("金额",),
            allow_aggregate=True,
            allow_filter=True,
        ),
        StructuredColumnSchema(
            physical_name="region",
            original_name="地区",
            display_name="地区",
            data_type=StructuredColumnType.STRING,
            aliases=("区域",),
            allow_aggregate=False,
            allow_filter=True,
        ),
        StructuredColumnSchema(
            physical_name="order_date",
            original_name="订单日期",
            display_name="订单日期",
            data_type=StructuredColumnType.DATE,
            aliases=("日期",),
            allow_aggregate=False,
            allow_filter=True,
        ),
    )


def sample_confirmed_schema(temp_dir: Path, row_count: int = 3) -> ConfirmedSpreadsheetFixture:
    rows: list[list[Any]] = [["订单金额", "地区", "订单日期"]]
    regions = ("华东", "华南", "华北")
    for index in range(row_count):
        rows.append(
            [str((index + 1) * 10), regions[index % len(regions)], f"2026-01-{index + 1:02d}"]
        )
    path = write_xlsx(temp_dir / "sales.xlsx", "明细", rows)
    schema = StructuredDatasetSchema(
        dataset_id="ds-sales",
        source_id="kb-sales",
        worksheet_name="明细",
        schema_version=1,
        columns=sample_columns(),
        schema_hash="a" * 64,
    )
    return ConfirmedSpreadsheetFixture(path=path, schema=schema)


def sample_publication() -> StructuredPublication:
    return StructuredPublication(
        publication_id="pub-sales-1",
        dataset_id="ds-sales",
        schema_version=1,
        physical_table_name="structured_ds_sales_v1",
        row_count=3,
        content_hash="b" * 64,
    )


def sample_catalog(*, ambiguous: bool = False) -> StructuredCatalog:
    columns = sample_columns()
    if ambiguous:
        columns += (
            StructuredColumnSchema(
                physical_name="net_amount",
                original_name="净金额",
                display_name="净金额",
                data_type=StructuredColumnType.DECIMAL,
                aliases=("金额",),
                allow_aggregate=True,
                allow_filter=True,
            ),
        )
    schema = StructuredDatasetSchema(
        dataset_id="ds-sales",
        source_id="kb-sales",
        worksheet_name="明细",
        schema_version=1,
        columns=columns,
        schema_hash="a" * 64,
    )
    return StructuredCatalog(
        datasets=(
            StructuredDatasetCatalog(
                schema=schema,
                source_name="sales.xlsx",
                active_publication=sample_publication(),
            ),
        ),
    )


class RecordingParquetSink:
    def __init__(self, root: Path | None = None) -> None:
        self.root = root or Path(".")
        self.batch_rows: list[int] = []
        self.output_paths: list[Path] = []
        self.batches: list[Any] = []

    def write_batch(self, rows: Sequence[Any], output_path: Path) -> None:
        self.batch_rows.append(len(rows))
        self.output_paths.append(Path(output_path))
        self.batches.append(rows)

    def write(self, rows: Sequence[Any], output_path: Path) -> None:
        self.write_batch(rows, output_path)

    def iter_batches(self, paths: Sequence[Path]):
        expected = [Path(path) for path in paths]
        selected = [
            batch
            for path, batch in zip(self.output_paths, self.batches, strict=True)
            if path in expected
        ]
        return iter(selected)


class FakeClickHouse:
    def __init__(self, aggregate_rows: Sequence[Any] = ()) -> None:
        self.ddl: list[Any] = []
        self.inserts: list[Any] = []
        self.queries: list[Any] = []
        self.aggregate_rows = list(aggregate_rows)

    def execute_ddl(self, statement: Any, *args: Any, **kwargs: Any) -> None:
        self.ddl.append(statement)

    def insert(self, *args: Any, **kwargs: Any) -> None:
        self.inserts.append((args, kwargs))

    def query(self, statement: Any, *args: Any, **kwargs: Any) -> list[Any]:
        self.queries.append((statement, args, kwargs))
        return list(self.aggregate_rows)

    def execute(self, statement: Any, *args: Any, **kwargs: Any) -> list[Any]:
        return self.query(statement, *args, **kwargs)


class RecordingLLMProvider:
    def __init__(self) -> None:
        self.generation_calls = 0
        self.requests: list[Any] = []

    def generate(self, request: Any, *args: Any, **kwargs: Any) -> Any:
        self.generation_calls += 1
        self.requests.append(request)
        return None

    def generate_reply(self, request: Any, *args: Any, **kwargs: Any) -> Any:
        return self.generate(request, *args, **kwargs)
