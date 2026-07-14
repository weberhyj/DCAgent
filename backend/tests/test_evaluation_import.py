from __future__ import annotations

import asyncio
import io
import inspect
import json
import time
import unittest
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from tempfile import TemporaryDirectory
from threading import Barrier, Event
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import event, select, text

from app.database import Database, EvaluationCaseRecord
from app.evaluation import evaluation_case_dedup_key, evaluation_case_lookup_keys
from app.evaluation_import import (
    MAX_IMPORT_BYTES,
    EvaluationImportFileError,
    EvaluationImportRow,
    EvaluationImportService,
    EvaluationImportTokenError,
    parse_boolean,
    read_import_records,
)
from app.main import create_app
from app.models import ChatState, KnowledgeSourceModel
from app.repository import InMemoryChatRepository
from app.routes import preview_evaluation_import
from app.schemas import EvaluationImportConfirmRequest
from app.sql_repository import SqlChatRepository


def make_source(source_id: str, name: str) -> KnowledgeSourceModel:
    return KnowledgeSourceModel(
        id=source_id,
        name=name,
        source_type="PDF",
        records=1,
        status="已索引",
        updated_at="2026-07-13 10:00:00",
        classification="内部",
    )


def json_content_at_size(size: int) -> bytes:
    prefix = b'[{"question":"q","expect_answer":false,"padding":"'
    suffix = b'"}]'
    padding_size = size - len(prefix) - len(suffix)
    if padding_size < 0:
        raise ValueError("size is too small for the JSON fixture")
    return prefix + (b"x" * padding_size) + suffix


def xlsx_content(headers: list[str] | None = None) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    if headers is not None:
        worksheet.append(headers)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class FailOnceEvaluationImportRepository:
    def __init__(self, repository: InMemoryChatRepository) -> None:
        self._repository = repository
        self._should_fail = True

    def __getattr__(self, name: str):
        return getattr(self._repository, name)

    def create_evaluation_cases(self, *args, **kwargs):
        if self._should_fail:
            self._should_fail = False
            raise RuntimeError("temporary repository failure")
        return self._repository.create_evaluation_cases(*args, **kwargs)


class RecordingUploadFile:
    def __init__(self, file_name: str, content: bytes) -> None:
        self.filename = file_name
        self._content = content
        self.read_sizes: list[int] = []

    async def read(self, size: int = -1) -> bytes:
        self.read_sizes.append(size)
        return self._content[:size] if size >= 0 else self._content


class EvaluationImportServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.service = EvaluationImportService(ttl_seconds=1800)
        self.sources = [make_source("kb-policy", "制度汇编.pdf")]

    def test_previews_utf8_bom_csv_answer_and_no_answer_rows(self) -> None:
        content = (
            "question,expect_answer,expected_sources,expected_terms,category,tags,top_k,external_key\n"
            "资料如何归档,true,制度汇编.pdf,归档|保管,制度,归档|流程,5,policy-001\n"
            "是否提供火星补贴,false,,,福利,无答案,3,no-answer-001\n"
        ).encode("utf-8-sig")

        preview = self.service.preview("cases.csv", content, self.sources)

        self.assertEqual(preview.total_rows, 2)
        self.assertEqual(preview.valid_rows, 2)
        self.assertEqual(preview.invalid_rows, 0)
        answer_row, no_answer_row = preview.rows
        self.assertEqual(answer_row.row_number, 2)
        self.assertEqual(answer_row.expected_source_ids, ["kb-policy"])
        self.assertEqual(answer_row.expected_terms, ["归档", "保管"])
        self.assertEqual(answer_row.category, "制度")
        self.assertEqual(answer_row.tags, ["归档", "流程"])
        self.assertEqual(answer_row.top_k, 5)
        self.assertEqual(answer_row.external_key, "policy-001")
        self.assertFalse(no_answer_row.expect_answer)
        self.assertEqual(no_answer_row.expected_source_ids, [])
        self.assertEqual(no_answer_row.expected_terms, [])

    def test_reports_json_row_validation_errors_with_file_row_numbers(self) -> None:
        content = json.dumps(
            [
                {"question": "", "expect_answer": True, "expected_terms": []},
                {
                    "question": "正常问题",
                    "expect_answer": True,
                    "expected_sources": ["不存在.pdf"],
                },
            ],
            ensure_ascii=False,
        ).encode("utf-8")

        preview = self.service.preview("cases.json", content, self.sources)

        self.assertEqual(preview.total_rows, 2)
        self.assertEqual(preview.valid_rows, 0)
        self.assertEqual(preview.invalid_rows, 2)
        self.assertEqual([error.row_number for error in preview.errors], [2, 3])
        self.assertIn("问题不能为空", preview.errors[0].message)
        self.assertIn("未找到资料", preview.errors[1].message)

    def test_reads_only_first_xlsx_worksheet(self) -> None:
        workbook = Workbook()
        first = workbook.active
        first.title = "导入"
        first.append(["question", "expect_answer", "expected_terms"])
        first.append(["第一张表的问题", "true", "关键词"])
        second = workbook.create_sheet("忽略")
        second.append(["question", "expect_answer", "expected_terms"])
        second.append(["第二张表的问题", "true", "另一个关键词"])
        buffer = io.BytesIO()
        workbook.save(buffer)

        preview = self.service.preview("cases.xlsx", buffer.getvalue(), self.sources)

        self.assertEqual(preview.total_rows, 1)
        self.assertEqual(preview.rows[0].question, "第一张表的问题")

    def test_rejects_xlsx_with_virtual_height_over_row_limit(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["question", "expect_answer"])
        worksheet.cell(row=3002, column=1, value="虚高末行")
        worksheet.cell(row=3002, column=2, value="false")
        buffer = io.BytesIO()
        workbook.save(buffer)

        with self.assertRaisesRegex(EvaluationImportFileError, "2000"):
            read_import_records("virtual-height.xlsx", buffer.getvalue())

    def test_xlsx_stops_reading_after_max_rows_plus_one(self) -> None:
        class GuardedWorksheet:
            def __init__(self) -> None:
                self.read_count = 0

            def iter_rows(self, values_only: bool = False):
                self.assert_values_only(values_only)
                yield ("question", "expect_answer")
                while self.read_count < 2001:
                    self.read_count += 1
                    yield ("q", "false")
                raise AssertionError("XLSX parser read past MAX_IMPORT_ROWS + 1")

            @staticmethod
            def assert_values_only(values_only: bool) -> None:
                if not values_only:
                    raise AssertionError("XLSX parser must request cell values only")

        class FakeWorkbook:
            def __init__(self, worksheet: GuardedWorksheet) -> None:
                self.worksheets = [worksheet]

            def close(self) -> None:
                return None

        worksheet = GuardedWorksheet()
        workbook = FakeWorkbook(worksheet)
        with (
            patch("app.evaluation_import._validate_xlsx_archive"),
            patch("app.evaluation_import.load_workbook", return_value=workbook),
        ):
            with self.assertRaisesRegex(EvaluationImportFileError, "2000"):
                read_import_records("guarded.xlsx", b"non-empty")
        self.assertEqual(worksheet.read_count, 2001)

    def test_rejects_xlsx_when_total_uncompressed_zip_entries_exceed_50mb(self) -> None:
        content = io.BytesIO(xlsx_content(["question", "expect_answer"]))
        with zipfile.ZipFile(content, mode="a", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("oversized-entry.bin", b"x" * (50 * 1024 * 1024 + 1))

        with self.assertRaisesRegex(EvaluationImportFileError, "解压后.*50MB"):
            read_import_records("inflated.xlsx", content.getvalue())

    def test_rejects_files_without_data_records(self) -> None:
        empty_record_files = [
            ("empty-array.json", b"[]"),
            ("header-only.csv", b"question,expect_answer\n"),
            ("empty-sheet.xlsx", xlsx_content()),
            ("header-only.xlsx", xlsx_content(["question", "expect_answer"])),
        ]

        for file_name, content in empty_record_files:
            with self.subTest(file_name=file_name):
                with self.assertRaisesRegex(EvaluationImportFileError, "无有效数据"):
                    read_import_records(file_name, content)

    def test_rejects_oversized_too_many_unsupported_and_empty_files(self) -> None:
        invalid_files = [
            ("oversized.csv", b"x" * (MAX_IMPORT_BYTES + 1), "5MB"),
            (
                "too-many.json",
                json.dumps([{"question": "q"}] * 2001).encode("utf-8"),
                "2000",
            ),
            ("cases.txt", b"question\nhello\n", "仅支持"),
            ("empty.csv", b"", "文件不能为空"),
        ]

        for file_name, content, message in invalid_files:
            with self.subTest(file_name=file_name):
                with self.assertRaisesRegex(EvaluationImportFileError, message):
                    read_import_records(file_name, content)

    def test_enforces_exact_file_size_boundary_with_valid_json(self) -> None:
        exact_limit = json_content_at_size(MAX_IMPORT_BYTES)
        over_limit = json_content_at_size(MAX_IMPORT_BYTES + 1)

        records = read_import_records("exact.json", exact_limit)

        self.assertEqual(len(exact_limit), MAX_IMPORT_BYTES)
        self.assertEqual(records[0]["question"], "q")
        with self.assertRaisesRegex(EvaluationImportFileError, "5MB"):
            read_import_records("over.json", over_limit)

    def test_enforces_exact_row_count_boundary(self) -> None:
        exact_limit = json.dumps(
            [{"question": f"q-{index}", "expect_answer": False} for index in range(2000)]
        ).encode("utf-8")
        over_limit = json.dumps(
            [{"question": f"q-{index}", "expect_answer": False} for index in range(2001)]
        ).encode("utf-8")

        preview = self.service.preview("exact.json", exact_limit, self.sources)

        self.assertEqual(preview.total_rows, 2000)
        self.assertEqual(preview.valid_rows, 2000)
        with self.assertRaisesRegex(EvaluationImportFileError, "2000"):
            read_import_records("over.json", over_limit)

    def test_rejects_large_csv_at_2001_rows(self) -> None:
        rows = ["question,expect_answer"]
        rows.extend(f"q-{index},false" for index in range(3002))

        with self.assertRaisesRegex(EvaluationImportFileError, "2000"):
            read_import_records("large.csv", ("\n".join(rows) + "\n").encode("utf-8"))

    def test_csv_stops_reading_after_max_rows_plus_one(self) -> None:
        class GuardedRows:
            def __init__(self) -> None:
                self.read_count = 0

            def __iter__(self) -> "GuardedRows":
                return self

            def __next__(self) -> dict[str, object]:
                self.read_count += 1
                if self.read_count <= 2001:
                    return {"question": "q", "expect_answer": "false"}
                raise AssertionError("CSV parser read past MAX_IMPORT_ROWS + 1")

        guarded_rows = GuardedRows()
        with patch("app.evaluation_import.csv.DictReader", return_value=guarded_rows):
            with self.assertRaisesRegex(EvaluationImportFileError, "2000"):
                read_import_records("guarded.csv", b"non-empty")
        self.assertEqual(guarded_rows.read_count, 2001)

    def test_converts_malformed_file_errors_to_chinese_file_errors(self) -> None:
        malformed_files = [
            ("bad-encoding.csv", b"question\n\xff\n", "UTF-8"),
            ("bad.json", b"[{", "JSON"),
            ("bad.xlsx", b"not-a-zip", "XLSX"),
            (
                "long-field.csv",
                b"question,expect_answer\n" + (b"x" * 200_000) + b",false\n",
                "CSV",
            ),
        ]

        for file_name, content, message in malformed_files:
            with self.subTest(file_name=file_name):
                with self.assertRaisesRegex(EvaluationImportFileError, message):
                    read_import_records(file_name, content)

    def test_parse_boolean_accepts_supported_values(self) -> None:
        truthy = [True, 1, "true", "TRUE", "1", "是", "应有答案"]
        falsy = [False, 0, "false", "FALSE", "0", "否", "应无答案"]

        for value in truthy:
            with self.subTest(value=value):
                self.assertTrue(parse_boolean(value))
        for value in falsy:
            with self.subTest(value=value):
                self.assertFalse(parse_boolean(value))

    def test_requires_explicit_non_empty_expect_answer(self) -> None:
        content = json.dumps(
            [
                {"question": "missing"},
                {"question": "empty", "expect_answer": ""},
            ]
        ).encode("utf-8")

        preview = self.service.preview("cases.json", content, self.sources)

        self.assertEqual(preview.invalid_rows, 2)
        self.assertEqual([error.field for error in preview.errors], ["expect_answer"] * 2)
        self.assertTrue(all("必须显式提供" in error.message for error in preview.errors))

    def test_rejects_container_values_for_scalar_and_list_fields(self) -> None:
        content = json.dumps(
            [
                {"question": ["bad"], "expect_answer": False},
                {"question": "category", "expect_answer": False, "category": ["bad"]},
                {
                    "question": "external key",
                    "expect_answer": False,
                    "external_key": {"bad": "value"},
                },
                {"question": "top k", "expect_answer": False, "top_k": [5]},
                {
                    "question": "sources",
                    "expect_answer": True,
                    "expected_sources": {"bad": "value"},
                },
                {
                    "question": "terms",
                    "expect_answer": True,
                    "expected_terms": {"bad": "value"},
                },
                {
                    "question": "tags",
                    "expect_answer": False,
                    "tags": {"bad": "value"},
                },
            ]
        ).encode("utf-8")

        preview = self.service.preview("malformed.json", content, self.sources)

        self.assertEqual(preview.valid_rows, 0)
        self.assertEqual(preview.invalid_rows, 7)
        self.assertEqual(
            [error.field for error in preview.errors],
            [
                "question",
                "category",
                "external_key",
                "top_k",
                "expected_sources",
                "expected_terms",
                "tags",
            ],
        )
        self.assertTrue(all("必须" in error.message for error in preview.errors))

    def test_reports_missing_and_ambiguous_source_names(self) -> None:
        sources = [
            make_source("kb-policy-1", "制度汇编.pdf"),
            make_source("kb-policy-2", " 制度汇编.pdf "),
        ]
        content = json.dumps(
            [
                {
                    "question": "同名资料",
                    "expect_answer": True,
                    "expected_sources": ["制度汇编.pdf"],
                },
                {
                    "question": "缺失资料",
                    "expect_answer": True,
                    "expected_sources": ["缺失.pdf"],
                },
            ],
            ensure_ascii=False,
        ).encode("utf-8")

        preview = self.service.preview("cases.json", content, sources)

        self.assertEqual(preview.invalid_rows, 2)
        self.assertIn("存在多个同名资料", preview.errors[0].message)
        self.assertIn("未找到资料", preview.errors[1].message)

    def test_validates_question_top_k_evidence_and_task1_metadata_limits(self) -> None:
        content = json.dumps(
            [
                {"question": "q" * 1001, "expect_answer": False},
                {"question": "top k", "expect_answer": False, "top_k": 11},
                {"question": "没有证据", "expect_answer": True},
                {
                    "question": "分类过长",
                    "expect_answer": False,
                    "category": "类" * 81,
                },
            ],
            ensure_ascii=False,
        ).encode("utf-8")

        preview = self.service.preview("cases.json", content, self.sources)

        self.assertEqual(preview.invalid_rows, 4)
        messages = [error.message for error in preview.errors]
        self.assertTrue(any("1000" in message for message in messages))
        self.assertTrue(any("1 到 10" in message for message in messages))
        self.assertTrue(any("必须提供期望资料或关键词" in message for message in messages))
        self.assertTrue(any("分类不能超过 80 个字符" in message for message in messages))

    def test_enforces_question_and_top_k_boundaries(self) -> None:
        content = json.dumps(
            [
                {"question": "q" * 1000, "expect_answer": False, "top_k": 1},
                {"question": "ten", "expect_answer": False, "top_k": 10},
                {"question": "q" * 1001, "expect_answer": False, "top_k": 5},
                {"question": "zero", "expect_answer": False, "top_k": 0},
                {"question": "eleven", "expect_answer": False, "top_k": 11},
            ]
        ).encode("utf-8")

        preview = self.service.preview("boundaries.json", content, self.sources)

        self.assertEqual(preview.valid_rows, 2)
        self.assertEqual([row.top_k for row in preview.rows], [1, 10])
        self.assertEqual(preview.invalid_rows, 3)
        self.assertEqual([error.row_number for error in preview.errors], [4, 5, 6])

    def test_reuses_task1_metadata_normalization_and_limits(self) -> None:
        twenty_tags = [f"tag-{index}" for index in range(20)]
        content = json.dumps(
            [
                {
                    "question": "normalization",
                    "expect_answer": False,
                    "category": "  制度  ",
                    "tags": ["归档", "归档", "流程"],
                    "external_key": "  case-001  ",
                },
                {
                    "question": "twenty tags",
                    "expect_answer": False,
                    "tags": twenty_tags,
                },
                {
                    "question": "tag length 80",
                    "expect_answer": False,
                    "tags": ["标" * 80],
                },
                {
                    "question": "key length 120",
                    "expect_answer": False,
                    "external_key": "k" * 120,
                },
                {
                    "question": "twenty one tags",
                    "expect_answer": False,
                    "tags": twenty_tags + ["extra"],
                },
                {
                    "question": "tag length 81",
                    "expect_answer": False,
                    "tags": ["标" * 81],
                },
                {
                    "question": "key length 121",
                    "expect_answer": False,
                    "external_key": "k" * 121,
                },
            ],
            ensure_ascii=False,
        ).encode("utf-8")

        preview = self.service.preview("metadata.json", content, self.sources)

        self.assertEqual(preview.valid_rows, 4)
        normalized = preview.rows[0]
        self.assertEqual(normalized.category, "制度")
        self.assertEqual(normalized.tags, ["归档", "流程"])
        self.assertEqual(normalized.external_key, "case-001")
        self.assertEqual(len(preview.rows[1].tags), 20)
        self.assertEqual(len(preview.rows[2].tags[0]), 80)
        self.assertEqual(len(preview.rows[3].external_key or ""), 120)
        self.assertEqual(preview.invalid_rows, 3)
        messages = [error.message for error in preview.errors]
        self.assertTrue(any("标签最多 20 个" in message for message in messages))
        self.assertTrue(any("单个标签不能超过 80 个字符" in message for message in messages))
        self.assertTrue(any("外部标识不能超过 120 个字符" in message for message in messages))

    def test_maps_task1_category_limit_error_to_stable_chinese(self) -> None:
        content = json.dumps(
            [
                {
                    "question": "category too long",
                    "expect_answer": False,
                    "category": "类" * 81,
                }
            ],
            ensure_ascii=False,
        ).encode("utf-8")

        preview = self.service.preview("metadata.json", content, self.sources)

        self.assertEqual(preview.invalid_rows, 1)
        self.assertEqual(preview.errors[0].field, "category")
        self.assertEqual(preview.errors[0].message, "分类不能超过 80 个字符")

    def test_consume_is_one_time_and_expired_tokens_have_chinese_error(self) -> None:
        content = json.dumps(
            [{"question": "无需答案", "expect_answer": False}],
            ensure_ascii=False,
        ).encode("utf-8")
        preview = self.service.preview("cases.json", content, self.sources)

        consumed = self.service.consume(preview.token)

        self.assertEqual(consumed.token, preview.token)
        with self.assertRaisesRegex(EvaluationImportTokenError, "导入预览已过期"):
            self.service.consume(preview.token)

        expired_service = EvaluationImportService(ttl_seconds=0)
        expired = expired_service.preview("cases.json", content, self.sources)
        with self.assertRaisesRegex(EvaluationImportTokenError, "请重新上传文件"):
            expired_service.consume(expired.token)

    def test_preview_token_uses_24_bytes_of_urlsafe_entropy(self) -> None:
        content = json.dumps(
            [{"question": "无需答案", "expect_answer": False}]
        ).encode("utf-8")

        with patch(
            "app.evaluation_import.secrets.token_urlsafe",
            return_value="preview-token",
        ) as token_urlsafe:
            preview = self.service.preview("cases.json", content, self.sources)

        self.assertEqual(preview.token, "preview-token")
        token_urlsafe.assert_called_once_with(24)

    def test_concurrent_consume_allows_exactly_one_success(self) -> None:
        content = json.dumps(
            [{"question": "无需答案", "expect_answer": False}]
        ).encode("utf-8")
        preview = self.service.preview("cases.json", content, self.sources)

        def consume() -> str:
            try:
                self.service.consume(preview.token)
                return "success"
            except EvaluationImportTokenError:
                return "expired"

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _: consume(), range(2)))

        self.assertEqual(sorted(results), ["expired", "success"])

    def test_concurrent_reserve_reports_confirming_and_release_allows_retry(self) -> None:
        content = json.dumps(
            [{"question": "可恢复令牌", "expect_answer": False}]
        ).encode("utf-8")
        preview = self.service.preview("cases.json", content, self.sources)
        self.assertTrue(hasattr(self.service, "reserve"))
        self.assertTrue(hasattr(self.service, "complete"))
        self.assertTrue(hasattr(self.service, "release"))
        barrier = Barrier(2)

        def reserve() -> str:
            barrier.wait()
            try:
                self.service.reserve(preview.token)
                return "reserved"
            except EvaluationImportTokenError as error:
                return str(error)

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(lambda _: reserve(), range(2)))

        self.assertEqual(sum(result == "reserved" for result in results), 1)
        self.assertEqual(sum("正在确认" in result for result in results), 1)

        self.service.release(preview.token)
        reserved_again = self.service.reserve(preview.token)
        self.assertEqual(reserved_again.token, preview.token)
        self.service.complete(preview.token)
        with self.assertRaisesRegex(EvaluationImportTokenError, "重新上传文件"):
            self.service.reserve(preview.token)

    def test_new_preview_cleans_expired_entries_before_capacity_check(self) -> None:
        service = EvaluationImportService(ttl_seconds=10, max_previews=1)
        content = json.dumps(
            [{"question": "无需答案", "expect_answer": False}]
        ).encode("utf-8")
        with patch("app.evaluation_import.time.time", return_value=100):
            expired = service.preview("first.json", content, self.sources)

        with patch("app.evaluation_import.time.time", return_value=111):
            current = service.preview("second.json", content, self.sources)
            with self.assertRaises(EvaluationImportTokenError):
                service.consume(expired.token)
            self.assertEqual(service.consume(current.token).file_name, "second.json")

    def test_rejects_preview_when_capacity_is_full(self) -> None:
        service = EvaluationImportService(ttl_seconds=1800, max_previews=1)
        content = json.dumps(
            [{"question": "无需答案", "expect_answer": False}]
        ).encode("utf-8")
        service.preview("first.json", content, self.sources)

        with self.assertRaisesRegex(EvaluationImportFileError, "预览数量已达上限"):
            service.preview("second.json", content, self.sources)

    def test_marks_duplicates_by_external_key_then_normalized_question(self) -> None:
        existing_cases = [
            SimpleNamespace(external_key="case-001", question="另一个问题"),
            SimpleNamespace(external_key=None, question="如何   归档？"),
        ]
        content = json.dumps(
            [
                {
                    "question": "新的问题",
                    "expect_answer": False,
                    "external_key": " case-001 ",
                },
                {"question": "  如何 归档？  ", "expect_answer": False},
                {
                    "question": "如何 归档？",
                    "expect_answer": False,
                    "external_key": "unique-key",
                },
            ],
            ensure_ascii=False,
        ).encode("utf-8")

        preview = self.service.preview(
            "cases.json",
            content,
            self.sources,
            existing_cases=existing_cases,
        )

        self.assertEqual(preview.valid_rows, 3)
        self.assertEqual(preview.duplicate_rows, 2)
        self.assertEqual(preview.duplicate_keys, ["case-001", "如何 归档？"])

    def test_casefolds_normalized_question_duplicate_keys(self) -> None:
        existing_cases = [
            SimpleNamespace(external_key=None, question="Hello   World"),
        ]
        content = json.dumps(
            [{"question": "  hello world  ", "expect_answer": False}]
        ).encode("utf-8")

        preview = self.service.preview(
            "cases.json",
            content,
            self.sources,
            existing_cases=existing_cases,
        )

        self.assertEqual(preview.duplicate_rows, 1)
        self.assertEqual(preview.duplicate_keys, ["hello world"])

    def test_dedup_helpers_normalize_blank_and_padded_external_keys(self) -> None:
        self.assertEqual(
            evaluation_case_dedup_key("  Normalize   Me ", " key "),
            "external_key:key",
        )
        self.assertEqual(
            evaluation_case_dedup_key("  Normalize   Me ", "   "),
            "question:normalize me",
        )
        self.assertEqual(
            evaluation_case_lookup_keys("  Normalize   Me ", " key "),
            {"external_key:key", "question:normalize me"},
        )
        self.assertEqual(
            evaluation_case_lookup_keys("  Normalize   Me ", "   "),
            {"question:normalize me"},
        )

    def test_materializes_existing_case_generator_once_for_duplicate_detection(self) -> None:
        existing_cases = (
            case
            for case in [SimpleNamespace(external_key=None, question="Generator Case")]
        )
        content = json.dumps(
            [{"question": "generator case", "expect_answer": False}]
        ).encode("utf-8")

        preview = self.service.preview(
            "cases.json",
            content,
            self.sources,
            existing_cases=existing_cases,
        )

        self.assertEqual(preview.duplicate_rows, 1)
        self.assertEqual(preview.duplicate_keys, ["generator case"])

    def test_existing_cases_default_is_none(self) -> None:
        parameter = inspect.signature(EvaluationImportService.preview).parameters[
            "existing_cases"
        ]

        self.assertIsNone(parameter.default)

    def test_create_app_initializes_import_service_with_1800_second_ttl(self) -> None:
        repository = InMemoryChatRepository(
            ChatState(conversations=[], messages_by_conversation={}, knowledge_sources=[])
        )

        app = create_app(repository=repository)

        service = app.state.evaluation_import_service
        self.assertIsInstance(service, EvaluationImportService)
        self.assertEqual(service.ttl_seconds, 1800)


class EvaluationImportApiTest(unittest.TestCase):
    def setUp(self) -> None:
        self.repository = InMemoryChatRepository(
            ChatState(
                conversations=[],
                messages_by_conversation={},
                knowledge_sources=[make_source("kb-policy", "制度汇编.pdf")],
            )
        )
        self.client = TestClient(create_app(repository=self.repository))

    def tearDown(self) -> None:
        self.client.close()

    def preview(self, rows: list[dict[str, object]], file_name: str = "cases.json"):
        return self.client.post(
            "/api/admin/evaluations/import/preview",
            files={
                "file": (
                    file_name,
                    json.dumps(rows, ensure_ascii=False).encode("utf-8"),
                    "application/json",
                )
            },
        )

    def test_preview_does_not_persist_and_confirm_persists_with_dashboard(self) -> None:
        response = self.preview(
            [
                {
                    "question": "资料如何归档",
                    "expect_answer": True,
                    "expected_sources": ["制度汇编.pdf"],
                    "expected_terms": ["归档"],
                    "category": "制度",
                    "tags": ["归档", "流程"],
                    "top_k": 3,
                    "external_key": "policy-001",
                },
                {"question": "", "expect_answer": False},
            ]
        )

        self.assertEqual(response.status_code, 200)
        preview = response.json()
        self.assertEqual(
            set(preview),
            {
                "previewToken",
                "fileName",
                "totalRows",
                "validRows",
                "invalidRows",
                "duplicateRows",
                "rows",
                "errors",
                "duplicateKeys",
            },
        )
        self.assertEqual(preview["fileName"], "cases.json")
        self.assertEqual(preview["totalRows"], 2)
        self.assertEqual(preview["validRows"], 1)
        self.assertEqual(preview["invalidRows"], 1)
        self.assertEqual(preview["duplicateRows"], 0)
        self.assertEqual(preview["rows"][0]["rowNumber"], 2)
        self.assertEqual(preview["rows"][0]["expectedSourceIds"], ["kb-policy"])
        self.assertEqual(preview["rows"][0]["externalKey"], "policy-001")
        self.assertEqual(preview["errors"][0]["rowNumber"], 3)
        self.assertEqual(preview["duplicateKeys"], [])
        self.assertEqual(self.repository.list_evaluation_cases(), [])

        confirm_response = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": preview["previewToken"]},
        )

        self.assertEqual(confirm_response.status_code, 200)
        confirmed = confirm_response.json()
        self.assertRegex(confirmed["importBatchId"], r"^eval-import-[0-9a-f]{12}$")
        self.assertEqual(confirmed["createdCount"], 1)
        self.assertEqual(confirmed["duplicateCount"], 0)
        self.assertEqual(len(confirmed["dashboard"]["cases"]), 1)
        case = self.repository.list_evaluation_cases()[0]
        self.assertEqual(case.import_batch_id, confirmed["importBatchId"])
        self.assertEqual(case.external_key, "policy-001")

    def test_preview_reads_only_max_bytes_plus_one(self) -> None:
        upload = RecordingUploadFile(
            "cases.json",
            json.dumps(
                [{"question": "限制读取", "expect_answer": False}],
                ensure_ascii=False,
            ).encode("utf-8"),
        )

        response = asyncio.run(
            preview_evaluation_import(
                file=upload,  # type: ignore[arg-type]
                repository=self.repository,
                service=self.client.app.state.evaluation_import_service,
            )
        )

        self.assertEqual(response.file_name, "cases.json")
        self.assertEqual(upload.read_sizes, [MAX_IMPORT_BYTES + 1])

    def test_preview_accepts_240_character_file_name_and_rejects_241_in_chinese(self) -> None:
        allowed_name = f"{'a' * 235}.json"
        rejected_name = f"{'b' * 236}.json"

        allowed = self.preview(
            [{"question": "文件名边界", "expect_answer": False}],
            file_name=allowed_name,
        )
        rejected = self.preview(
            [{"question": "文件名过长", "expect_answer": False}],
            file_name=rejected_name,
        )

        self.assertEqual(len(allowed_name), 240)
        self.assertEqual(allowed.status_code, 200)
        self.assertEqual(len(rejected_name), 241)
        self.assertEqual(rejected.status_code, 400)
        self.assertIn("文件名不能超过 240 个字符", rejected.json()["detail"])

    def test_preview_returns_row_errors_but_rejects_bad_files_and_no_valid_rows(self) -> None:
        partial = self.preview(
            [
                {"question": "可导入", "expect_answer": False},
                {"question": "", "expect_answer": False},
            ]
        )
        self.assertEqual(partial.status_code, 200)
        self.assertEqual(partial.json()["invalidRows"], 1)
        self.assertTrue(partial.json()["errors"])

        invalid_requests = [
            (
                {"file": ("cases.txt", b"question\nhello\n", "text/plain")},
                "仅支持",
            ),
            (
                {
                    "file": (
                        "large.csv",
                        b"x" * (MAX_IMPORT_BYTES + 1),
                        "text/csv",
                    )
                },
                "5MB",
            ),
            (
                {
                    "file": (
                        "invalid.json",
                        json.dumps(
                            [{"question": "", "expect_answer": False}],
                            ensure_ascii=False,
                        ).encode("utf-8"),
                        "application/json",
                    )
                },
                "没有可导入的有效数据",
            ),
        ]
        for files, message in invalid_requests:
            with self.subTest(message=message):
                response = self.client.post(
                    "/api/admin/evaluations/import/preview",
                    files=files,
                )
                self.assertEqual(response.status_code, 400)
                self.assertIn(message, response.json()["detail"])

    def test_confirm_token_is_single_use_with_consistent_chinese_error(self) -> None:
        preview = self.preview(
            [{"question": "一次性 token", "expect_answer": False}]
        ).json()

        first = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": preview["previewToken"]},
        )
        second = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": preview["previewToken"]},
        )

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 410)
        self.assertIn("预览", second.json()["detail"])
        self.assertIn("重新上传", second.json()["detail"])

    def test_confirm_releases_token_after_repository_failure_so_retry_can_succeed(self) -> None:
        repository = InMemoryChatRepository(
            ChatState(
                conversations=[],
                messages_by_conversation={},
                knowledge_sources=[make_source("kb-policy", "制度汇编.pdf")],
            )
        )
        wrapped_repository = FailOnceEvaluationImportRepository(repository)
        with TestClient(
            create_app(repository=wrapped_repository),
            raise_server_exceptions=False,
        ) as client:
            preview = client.post(
                "/api/admin/evaluations/import/preview",
                files={
                    "file": (
                        "retry.json",
                        json.dumps(
                            [{"question": "仓储失败后重试", "expect_answer": False}],
                            ensure_ascii=False,
                        ).encode("utf-8"),
                        "application/json",
                    )
                },
            ).json()
            payload = {"previewToken": preview["previewToken"]}

            first = client.post(
                "/api/admin/evaluations/import/confirm",
                json=payload,
            )
            second = client.post(
                "/api/admin/evaluations/import/confirm",
                json=payload,
            )

        self.assertEqual(first.status_code, 500)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.json()["createdCount"], 1)
        self.assertEqual(len(repository.list_evaluation_cases()), 1)

    def test_manual_duplicate_returns_chinese_conflict(self) -> None:
        payload = {
            "question": "手工重复用例",
            "expectAnswer": False,
            "externalKey": "manual-duplicate-001",
        }

        first = self.client.post("/api/admin/evaluations/cases", json=payload)
        second = self.client.post("/api/admin/evaluations/cases", json=payload)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 409)
        self.assertIn("评测用例已存在", second.json()["detail"])
        self.assertEqual(len(self.repository.list_evaluation_cases()), 1)

    def test_confirm_expired_token_uses_the_same_chinese_gone_response(self) -> None:
        self.client.app.state.evaluation_import_service = EvaluationImportService(
            ttl_seconds=0
        )
        preview = self.preview(
            [{"question": "立即过期", "expect_answer": False}]
        ).json()

        response = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": preview["previewToken"]},
        )

        self.assertEqual(response.status_code, 410)
        self.assertIn("预览", response.json()["detail"])
        self.assertIn("重新上传", response.json()["detail"])

    def test_confirm_rejects_missing_null_empty_and_blank_preview_tokens_in_chinese(self) -> None:
        invalid_payloads = [
            {},
            {"previewToken": None},
            {"previewToken": ""},
            {"previewToken": "   "},
        ]

        for payload in invalid_payloads:
            with self.subTest(payload=payload):
                response = self.client.post(
                    "/api/admin/evaluations/import/confirm",
                    json=payload,
                )

                self.assertEqual(response.status_code, 422)
                detail = json.dumps(response.json()["detail"], ensure_ascii=False)
                self.assertIn("预览令牌不能为空", detail)

    def test_confirm_request_openapi_requires_preview_token_without_default(self) -> None:
        schema = EvaluationImportConfirmRequest.model_json_schema(by_alias=True)

        self.assertIn("previewToken", schema["required"])
        self.assertNotIn("default", schema["properties"]["previewToken"])

    def test_confirm_rechecks_duplicates_after_preview(self) -> None:
        preview = self.preview(
            [
                {
                    "question": "竞态问题",
                    "expect_answer": False,
                    "external_key": "race-001",
                }
            ]
        ).json()
        self.repository.create_evaluation_case(
            question="其他问题",
            expected_source_ids=[],
            expected_terms=[],
            top_k=5,
            expect_answer=False,
            external_key="race-001",
        )

        response = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": preview["previewToken"]},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["createdCount"], 0)
        self.assertEqual(response.json()["duplicateCount"], 1)
        self.assertEqual(len(self.repository.list_evaluation_cases()), 1)

    def test_reimport_reports_duplicates_and_creates_no_second_copy(self) -> None:
        rows = [
            {
                "question": "按外部标识去重",
                "expect_answer": False,
                "external_key": "repeat-001",
            },
            {"question": "  Normalize   Me  ", "expect_answer": False},
        ]
        first_preview = self.preview(rows).json()
        first_confirm = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": first_preview["previewToken"]},
        )
        second_preview_response = self.preview(rows)

        self.assertEqual(first_confirm.status_code, 200)
        self.assertEqual(first_confirm.json()["createdCount"], 2)
        self.assertEqual(second_preview_response.status_code, 200)
        second_preview = second_preview_response.json()
        self.assertEqual(second_preview["duplicateRows"], 2)
        self.assertEqual(second_preview["duplicateKeys"], ["repeat-001", "normalize me"])

        second_confirm = self.client.post(
            "/api/admin/evaluations/import/confirm",
            json={"previewToken": second_preview["previewToken"]},
        )
        self.assertEqual(second_confirm.status_code, 200)
        self.assertEqual(second_confirm.json()["createdCount"], 0)
        self.assertEqual(second_confirm.json()["duplicateCount"], 2)
        self.assertEqual(len(self.repository.list_evaluation_cases()), 2)


class EvaluationImportSqlRepositoryTest(unittest.TestCase):
    def setUp(self) -> None:
        self.database = Database("sqlite+pysqlite:///:memory:")
        self.database.create_schema()
        self.repository = SqlChatRepository(self.database)

    def tearDown(self) -> None:
        self.database.engine.dispose()

    @staticmethod
    def rows() -> list[EvaluationImportRow]:
        return [
            EvaluationImportRow(
                row_number=2,
                question="第一行",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category="回归",
                tags=["批量"],
                top_k=5,
                external_key="sql-001",
            ),
            EvaluationImportRow(
                row_number=3,
                question="第二行",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=3,
                external_key=None,
            ),
        ]

    def test_import_batch_and_cases_persist_across_repository_rebuild(self) -> None:
        result = self.repository.create_evaluation_cases(
            rows=self.rows(),
            import_batch_id="eval-import-persist1",
            file_name="cases.json",
            total_rows=3,
            valid_rows=2,
            invalid_rows=1,
        )

        rebuilt = SqlChatRepository(self.database)
        cases = rebuilt.list_evaluation_cases()
        batches = rebuilt.list_evaluation_import_batches()

        self.assertEqual(result.created_count, 2)
        self.assertEqual(result.duplicate_count, 0)
        self.assertEqual(len(cases), 2)
        self.assertTrue(all(case.import_batch_id == "eval-import-persist1" for case in cases))
        self.assertEqual(len(batches), 1)
        batch = batches[0]
        self.assertEqual(batch.id, "eval-import-persist1")
        self.assertEqual(batch.file_name, "cases.json")
        self.assertEqual(batch.status, "completed")
        self.assertEqual(batch.total_rows, 3)
        self.assertEqual(batch.valid_rows, 2)
        self.assertEqual(batch.invalid_rows, 1)
        self.assertEqual(batch.duplicate_rows, 0)
        self.assertTrue(batch.created_at)
        self.assertTrue(batch.completed_at)

    def test_database_marks_memory_sqlite_and_exposes_reentrant_write_lock(self) -> None:
        self.assertTrue(self.database.is_sqlite_memory)
        with self.database.write_lock:
            with self.database.write_lock:
                pass

    def test_create_evaluation_cases_signature_drops_preview_duplicate_count(self) -> None:
        for method in (
            InMemoryChatRepository.create_evaluation_cases,
            SqlChatRepository.create_evaluation_cases,
        ):
            with self.subTest(method=method.__qualname__):
                self.assertNotIn(
                    "duplicate_rows",
                    inspect.signature(method).parameters,
                )

    def test_concurrent_memory_sqlite_imports_share_database_lock(self) -> None:
        repositories = [
            SqlChatRepository(self.database),
            SqlChatRepository(self.database),
        ]
        barrier = Barrier(2)

        def confirm(worker: int):
            barrier.wait()
            return repositories[worker].create_evaluation_cases(
                rows=[
                    EvaluationImportRow(
                        row_number=2,
                        question=f"内存并发问题 {worker}",
                        expect_answer=False,
                        expected_source_ids=[],
                        expected_terms=[],
                        category=None,
                        tags=[],
                        top_k=5,
                        external_key="memory-concurrent-key",
                    )
                ],
                import_batch_id=f"eval-import-memory-{worker}",
                file_name="memory.json",
                total_rows=1,
                valid_rows=1,
                invalid_rows=0,
            )

        with ThreadPoolExecutor(max_workers=2) as executor:
            futures = [executor.submit(confirm, worker) for worker in range(2)]
            results = [future.result() for future in futures]

        self.assertEqual(len(self.repository.list_evaluation_cases()), 1)
        self.assertEqual(sum(result.created_count for result in results), 1)
        self.assertEqual(sum(result.duplicate_count for result in results), 1)

    def test_memory_sqlite_read_session_cannot_commit_paused_failed_import(self) -> None:
        existing_case = self.repository.create_evaluation_case(
            question="已有用例",
            expected_source_ids=[],
            expected_terms=[],
            top_k=5,
            expect_answer=False,
            external_key="existing-case",
        )
        with self.database.engine.begin() as connection:
            connection.execute(
                text(
                    """
                    CREATE TRIGGER fail_paused_import_second_case
                    BEFORE INSERT ON evaluation_cases
                    WHEN NEW.question = '失败的第二行'
                    BEGIN
                        SELECT RAISE(ABORT, 'paused second row failed');
                    END;
                    """
                )
            )

        update_paused = Event()
        continue_write = Event()
        read_started = Event()

        def pause_after_sort_order_update(
            connection,
            cursor,
            statement,
            parameters,
            context,
            executemany,
        ) -> None:
            normalized_statement = " ".join(statement.lower().split())
            if normalized_statement.startswith("update evaluation_cases set sort_order="):
                update_paused.set()
                if not continue_write.wait(timeout=5):
                    raise TimeoutError("timed out waiting to resume paused import")

        def import_cases():
            try:
                self.repository.create_evaluation_cases(
                    rows=[
                        EvaluationImportRow(
                            row_number=2,
                            question="成功的第一行",
                            expect_answer=False,
                            expected_source_ids=[],
                            expected_terms=[],
                            category=None,
                            tags=[],
                            top_k=5,
                            external_key="paused-first",
                        ),
                        EvaluationImportRow(
                            row_number=3,
                            question="失败的第二行",
                            expect_answer=False,
                            expected_source_ids=[],
                            expected_terms=[],
                            category=None,
                            tags=[],
                            top_k=5,
                            external_key="paused-second",
                        ),
                    ],
                    import_batch_id="eval-import-paused-failure",
                    file_name="paused.json",
                    total_rows=2,
                    valid_rows=2,
                    invalid_rows=0,
                )
            except Exception as error:
                return error
            raise AssertionError("import unexpectedly succeeded")

        def read_cases():
            read_started.set()
            return SqlChatRepository(self.database).list_evaluation_cases()

        event.listen(
            self.database.engine,
            "after_cursor_execute",
            pause_after_sort_order_update,
        )
        try:
            with ThreadPoolExecutor(max_workers=2) as executor:
                import_future = executor.submit(import_cases)
                self.assertTrue(update_paused.wait(timeout=5))
                read_future = executor.submit(read_cases)
                self.assertTrue(read_started.wait(timeout=5))
                time.sleep(0.1)
                continue_write.set()
                import_error = import_future.result(timeout=5)
                read_future.result(timeout=5)
        finally:
            continue_write.set()
            event.remove(
                self.database.engine,
                "after_cursor_execute",
                pause_after_sort_order_update,
            )

        self.assertIn("paused second row failed", str(import_error))
        with self.database.session() as session:
            existing_sort_order = session.scalar(
                select(EvaluationCaseRecord.sort_order).where(
                    EvaluationCaseRecord.id == existing_case.id
                )
            )
        self.assertEqual(existing_sort_order, 0)
        self.assertEqual(
            [case.id for case in self.repository.list_evaluation_cases()],
            [existing_case.id],
        )
        self.assertEqual(self.repository.list_evaluation_import_batches(), [])

    def test_direct_repositories_normalize_padded_external_key_for_batch_dedup(self) -> None:
        repositories = [
            InMemoryChatRepository(
                ChatState(
                    conversations=[],
                    messages_by_conversation={},
                    knowledge_sources=[],
                )
            ),
            self.repository,
        ]

        for index, repository in enumerate(repositories):
            with self.subTest(repository=type(repository).__name__):
                repository.create_evaluation_case(
                    question=f"手工创建 {index}",
                    expected_source_ids=[],
                    expected_terms=[],
                    top_k=5,
                    expect_answer=False,
                    external_key=" key ",
                )
                result = repository.create_evaluation_cases(
                    rows=[
                        EvaluationImportRow(
                            row_number=2,
                            question=f"批量导入 {index}",
                            expect_answer=False,
                            expected_source_ids=[],
                            expected_terms=[],
                            category=None,
                            tags=[],
                            top_k=5,
                            external_key=" key ",
                        )
                    ],
                    import_batch_id=f"eval-import-padded-key-{index}",
                    file_name="padded-key.json",
                    total_rows=1,
                    valid_rows=1,
                    invalid_rows=0,
                )

                self.assertEqual(result.created_count, 0)
                self.assertEqual(result.duplicate_count, 1)
                self.assertEqual(len(repository.list_evaluation_cases()), 1)

    def test_confirm_time_duplicate_check_uses_external_key_before_question(self) -> None:
        self.repository.create_evaluation_case(
            question="Existing   Question",
            expected_source_ids=[],
            expected_terms=[],
            top_k=5,
            expect_answer=False,
            external_key="existing-key",
        )
        rows = [
            EvaluationImportRow(
                row_number=2,
                question="不同问题",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=5,
                external_key="existing-key",
            ),
            EvaluationImportRow(
                row_number=3,
                question="  existing question  ",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=5,
                external_key=None,
            ),
            EvaluationImportRow(
                row_number=4,
                question="existing question",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=5,
                external_key="new-key",
            ),
        ]

        result = self.repository.create_evaluation_cases(
            rows=rows,
            import_batch_id="eval-import-dedup01",
            file_name="dedup.json",
            total_rows=3,
            valid_rows=3,
            invalid_rows=0,
        )

        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.duplicate_count, 2)
        cases = self.repository.list_evaluation_cases()
        self.assertEqual(len(cases), 2)
        self.assertEqual(cases[0].external_key, "new-key")
        self.assertEqual(
            self.repository.list_evaluation_import_batches()[0].duplicate_rows,
            2,
        )

    def test_concurrent_file_sqlite_imports_serialize_external_and_question_keys(self) -> None:
        scenarios = [
            (
                "external-key",
                lambda worker: EvaluationImportRow(
                    row_number=2,
                    question=f"并发外部标识问题 {worker}",
                    expect_answer=False,
                    expected_source_ids=[],
                    expected_terms=[],
                    category=None,
                    tags=[],
                    top_k=5,
                    external_key="concurrent-key",
                ),
            ),
            (
                "normalized-question",
                lambda worker: EvaluationImportRow(
                    row_number=2,
                    question=(
                        "Concurrent   Question"
                        if worker == 0
                        else "  concurrent question  "
                    ),
                    expect_answer=False,
                    expected_source_ids=[],
                    expected_terms=[],
                    category=None,
                    tags=[],
                    top_k=5,
                    external_key=None,
                ),
            ),
        ]

        with TemporaryDirectory() as temporary_directory:
            for scenario, row_factory in scenarios:
                with self.subTest(scenario=scenario):
                    database_path = Path(temporary_directory) / f"{scenario}.db"
                    database_url = f"sqlite+pysqlite:///{database_path.as_posix()}"
                    databases = [Database(database_url), Database(database_url)]
                    databases[0].create_schema()
                    repositories = [
                        SqlChatRepository(database) for database in databases
                    ]
                    barrier = Barrier(2)
                    verifier_database = None

                    def confirm(worker: int):
                        barrier.wait()
                        return repositories[worker].create_evaluation_cases(
                            rows=[row_factory(worker)],
                            import_batch_id=f"eval-import-race-{scenario}-{worker}",
                            file_name="race.json",
                            total_rows=1,
                            valid_rows=1,
                            invalid_rows=0,
                        )

                    try:
                        with ThreadPoolExecutor(max_workers=2) as executor:
                            results = list(executor.map(confirm, range(2)))

                        verifier_database = Database(database_url)
                        verifier = SqlChatRepository(verifier_database)
                        cases = verifier.list_evaluation_cases()
                        self.assertEqual(len(cases), 1)
                        self.assertEqual(
                            sum(result.created_count for result in results),
                            1,
                        )
                        self.assertEqual(
                            sum(result.duplicate_count for result in results),
                            1,
                        )
                    finally:
                        if verifier_database is not None:
                            verifier_database.engine.dispose()
                        for database in databases:
                            database.engine.dispose()

    def test_file_sqlite_manual_create_and_batch_import_serialize_dedup(self) -> None:
        scenarios = [
            (
                "external-key",
                "手工外部标识问题",
                "  批量外部标识问题  ",
                "manual-import-race-key",
            ),
            (
                "normalized-question",
                "Manual   Import Question",
                "  manual import question  ",
                None,
            ),
        ]

        with TemporaryDirectory() as temporary_directory:
            for scenario, manual_question, import_question, external_key in scenarios:
                with self.subTest(scenario=scenario):
                    database_path = Path(temporary_directory) / f"manual-{scenario}.db"
                    database_url = f"sqlite+pysqlite:///{database_path.as_posix()}"
                    import_database = Database(database_url)
                    manual_database = Database(database_url)
                    import_database.create_schema()
                    import_repository = SqlChatRepository(import_database)
                    manual_repository = SqlChatRepository(manual_database)
                    barrier = Barrier(2)

                    def import_cases():
                        barrier.wait()
                        return import_repository.create_evaluation_cases(
                            rows=[
                                EvaluationImportRow(
                                    row_number=2,
                                    question=import_question,
                                    expect_answer=False,
                                    expected_source_ids=[],
                                    expected_terms=[],
                                    category=None,
                                    tags=[],
                                    top_k=5,
                                    external_key=external_key,
                                )
                            ],
                            import_batch_id=f"eval-import-manual-race-{scenario}",
                            file_name="race.json",
                            total_rows=1,
                            valid_rows=1,
                            invalid_rows=0,
                        )

                    def create_manually():
                        barrier.wait()
                        time.sleep(0.05)
                        try:
                            return manual_repository.create_evaluation_case(
                                question=manual_question,
                                expected_source_ids=[],
                                expected_terms=[],
                                top_k=5,
                                expect_answer=False,
                                external_key=external_key,
                            )
                        except Exception as error:
                            return error

                    try:
                        with ThreadPoolExecutor(max_workers=2) as executor:
                            results = list(
                                executor.map(
                                    lambda operation: operation(),
                                    (import_cases, create_manually),
                                )
                            )

                        verifier = SqlChatRepository(import_database)
                        self.assertEqual(len(verifier.list_evaluation_cases()), 1)
                        self.assertFalse(
                            any("OperationalError" in type(result).__name__ for result in results)
                        )
                    finally:
                        manual_database.engine.dispose()
                        import_database.engine.dispose()

    def test_postgres_import_locks_use_sorted_database_hashed_dedup_keys(self) -> None:
        session = MagicMock()
        session.get_bind.return_value = SimpleNamespace(
            dialect=SimpleNamespace(name="postgresql")
        )
        rows = [
            EvaluationImportRow(
                row_number=2,
                question="Second question",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=5,
                external_key="key-b",
            ),
            EvaluationImportRow(
                row_number=3,
                question="  Normalize   Me ",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=5,
                external_key=None,
            ),
            EvaluationImportRow(
                row_number=4,
                question="First question",
                expect_answer=False,
                expected_source_ids=[],
                expected_terms=[],
                category=None,
                tags=[],
                top_k=5,
                external_key="key-a",
            ),
        ]

        self.repository._lock_evaluation_case_dedup_keys(
            session,
            [
                "external_key:key-b",
                "question:normalize me",
                "external_key:key-a",
            ],
        )

        lock_calls = session.execute.call_args_list
        self.assertEqual(
            [call.args[1]["dedup_key"] for call in lock_calls],
            [
                "external_key:key-a",
                "external_key:key-b",
                "question:normalize me",
            ],
        )
        for call in lock_calls:
            statement = str(call.args[0])
            self.assertIn("pg_advisory_xact_lock", statement)
            self.assertIn("hashtextextended", statement)

    def test_second_case_insert_failure_rolls_back_batch_and_all_cases(self) -> None:
        with self.database.engine.begin() as connection:
            connection.execute(
                text(
                    """
                    CREATE TRIGGER fail_second_evaluation_case
                    BEFORE INSERT ON evaluation_cases
                    WHEN NEW.question = '第二行'
                    BEGIN
                        SELECT RAISE(ABORT, 'second row failed');
                    END;
                    """
                )
            )

        with self.assertRaisesRegex(Exception, "second row failed"):
            self.repository.create_evaluation_cases(
                rows=self.rows(),
                import_batch_id="eval-import-rollback",
                file_name="cases.json",
                total_rows=2,
                valid_rows=2,
                invalid_rows=0,
            )

        rebuilt = SqlChatRepository(self.database)
        self.assertEqual(rebuilt.list_evaluation_cases(), [])
        self.assertEqual(rebuilt.list_evaluation_import_batches(), [])


if __name__ == "__main__":
    unittest.main()
