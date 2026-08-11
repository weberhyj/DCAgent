from __future__ import annotations

import os
import re
import tempfile
import unittest
import uuid
from dataclasses import replace
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.clickhouse_gateway import ClickHouseGateway
from app.database import Database
from app.main import create_app
from app.models import ChatMessageModel, ResponseParagraphModel
from app.offline_settings import require_secret_file
from app.sql_repository import SqlChatRepository
from app.structured_answer import StructuredAnswerService
from app.structured_ingestion import ArrowParquetSink, SpreadsheetPublisher
from app.structured_repository import StructuredRepository
from app.structured_worker import StructuredIngestionWorker
from tests.support.structured_fakes import sample_catalog, sample_multi_metric_catalog

ROW_COUNT = 100_000
MULTI_ROW_COUNT = 100_001
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _fixture_row(index: int) -> tuple[Decimal | None, str, date]:
    amount = None if index % 97 == 0 else Decimal(index % 100_000) / Decimal(100)
    region = ("华东", "华南", "华北", "西部")[index % 4]
    order_date = date(2025, 1, 1) + timedelta(days=index % 365)
    return amount, region, order_date


def _write_fixture(path: Path, rows: int = ROW_COUNT) -> Path:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Sales")
    sheet.append(("order_amount", "region", "order_date"))
    for index in range(rows):
        sheet.append(_fixture_row(index))
    workbook.save(path)
    workbook.close()
    return path


def _multi_fixture_row(index: int) -> tuple[str, Decimal | None, Decimal | None, Decimal | None]:
    region = ("华东", "华南", "华北", "西部")[index % 4]
    sales = None if index % 113 == 0 else Decimal((index * 17) % 100_000) / Decimal(100)
    cost = None if index % 127 == 0 else Decimal((index * 11 + 7) % 80_000) / Decimal(100)
    profit = None if index % 149 == 0 else Decimal((index * 5 - 13) % 40_000) / Decimal(100)
    return region, sales, cost, profit


def _write_multi_fixture(path: Path, rows: int = MULTI_ROW_COUNT) -> Path:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("明细")
    sheet.append(("地区", "销售额", "成本", "利润"))
    for index in range(rows):
        sheet.append(_multi_fixture_row(index))
    workbook.save(path)
    workbook.close()
    return path


def _multi_decimal_reference(region: str) -> dict[str, Decimal | int | None]:
    names = ("sales_amount", "cost_amount", "profit_amount")
    sums = {name: Decimal(0) for name in names}
    valid_counts = {name: 0 for name in names}
    total_count = 0
    for index in range(MULTI_ROW_COUNT):
        row_region, sales, cost, profit = _multi_fixture_row(index)
        if row_region != region:
            continue
        total_count += 1
        for name, value in zip(names, (sales, cost, profit), strict=True):
            if value is not None:
                sums[name] += value
                valid_counts[name] += 1

    result: dict[str, Decimal | int | None] = {"total_count": total_count}
    for name in names:
        result[f"{name}_value"] = sums[name] if valid_counts[name] else None
        result[f"{name}_valid_count"] = valid_counts[name]
        result[f"{name}_null_count"] = total_count - valid_counts[name]
    return result


class PhysocFake:
    def __init__(self) -> None:
        self.calls = 0

    def generate_reply(self, _request: object) -> ChatMessageModel:
        self.calls += 1
        return ChatMessageModel(
            id=f"msg-physoc-{self.calls}",
            role="assistant",
            time="2026-07-23 12:00:00",
            paragraphs=[ResponseParagraphModel(text="legacy Physoc answer")],
        )


class SwitchableCatalog:
    def __init__(self, repository: StructuredRepository) -> None:
        self.repository = repository
        self.override = None

    def __call__(self):
        return self.override or self.repository.get_catalog()


class InMemoryAggregateGateway:
    def __init__(self) -> None:
        self.rows: list[tuple[Decimal | None, str, date]] = []
        self.query_calls: list[tuple[str, dict[str, object]]] = []
        self.last_result: dict[str, object] | None = None
        self.error: Exception | None = None
        self.batch_rows: list[int] = []

    def prepare_publication(
        self,
        _schema: object,
        _publication_id: str,
        content_hash: str,
        **_kwargs: object,
    ) -> object:
        self.rows.clear()
        self.batch_rows.clear()
        return SimpleNamespace(
            physical_table_name="structured_e2e_sales",
            staging_table="structured_e2e_sales_staging",
            content_hash=content_hash,
        )

    def insert_batch(self, _target: object, batch: object) -> None:
        amount_index = batch.schema.get_field_index("order_amount")
        region_index = batch.schema.get_field_index("region")
        date_index = batch.schema.get_field_index("order_date")
        amounts = batch.column(amount_index).to_pylist()
        regions = batch.column(region_index).to_pylist()
        dates = batch.column(date_index).to_pylist()
        self.rows.extend(zip(amounts, regions, dates, strict=True))
        self.batch_rows.append(batch.num_rows)

    def validate_and_promote(self, target: object, **statistics: object) -> str:
        if statistics["row_count"] != len(self.rows):
            raise AssertionError("published row count does not match inserted rows")
        return target.physical_table_name

    def discard_publication(self, _target: object) -> None:
        self.rows.clear()

    def query(self, statement: str, parameters: object) -> dict[str, object]:
        if self.error is not None:
            raise self.error
        bound = dict(parameters)
        self.query_calls.append((statement, bound))
        filtered = [row for row in self.rows if self._matches(statement, bound, row)]
        match = re.search(
            r"SELECT\s+(avg|sum|count|min|max)\(([^)]*)\)\s+AS aggregate_value",
            statement,
            re.IGNORECASE,
        )
        if match is None:
            raise AssertionError(statement)
        aggregate = match.group(1).lower()
        metric = match.group(2).strip()
        values = [row[0] for row in filtered if row[0] is not None]
        if aggregate == "count":
            aggregate_value: Decimal | int | None = len(filtered) if not metric else len(values)
        elif not values:
            aggregate_value = None
        elif aggregate == "avg":
            aggregate_value = sum(values, Decimal(0)) / len(values)
        elif aggregate == "sum":
            aggregate_value = sum(values, Decimal(0))
        elif aggregate == "min":
            aggregate_value = min(values)
        else:
            aggregate_value = max(values)
        self.last_result = {
            "aggregate_value": aggregate_value,
            "total_count": len(filtered),
            "valid_count": len(filtered) if not metric else len(values),
            "null_count": 0 if not metric else len(filtered) - len(values),
        }
        return self.last_result

    @staticmethod
    def _matches(
        statement: str,
        parameters: dict[str, object],
        row: tuple[Decimal | None, str, date],
    ) -> bool:
        values = {"order_amount": row[0], "region": row[1], "order_date": row[2]}
        for name, expected in parameters.items():
            pattern = re.compile(
                rf"([a-z0-9_]+)\s*(>=|<=|=|>|<)\s*\{{{re.escape(name)}:[^}}]+\}}",
                re.IGNORECASE,
            )
            match = pattern.search(statement)
            if match is None:
                raise AssertionError(f"parameter {name} is not bound in SQL: {statement}")
            actual = values[match.group(1)]
            if actual is None:
                return False
            operator = match.group(2)
            if operator == "=" and actual != expected:
                return False
            if operator == ">=" and actual < expected:
                return False
            if operator == "<=" and actual > expected:
                return False
            if operator == ">" and actual <= expected:
                return False
            if operator == "<" and actual >= expected:
                return False
        return True

    def close(self) -> None:
        return None


class LargeMultiSummaryGateway:
    """Ingestion/query fake that never materializes source rows for query evaluation."""

    def __init__(self, result: dict[str, object]) -> None:
        self.result = result
        self.query_calls = 0
        self.returned_row_count = 0
        self.inserted_row_count = 0
        self.batch_rows: list[int] = []
        self.statements: list[str] = []
        self.fail_with: Exception | None = None

    def prepare_publication(
        self,
        _schema: object,
        _publication_id: str,
        content_hash: str,
        **_kwargs: object,
    ) -> object:
        self.inserted_row_count = 0
        self.batch_rows.clear()
        return SimpleNamespace(
            physical_table_name="structured_e2e_large_multi",
            staging_table="structured_e2e_large_multi_staging",
            content_hash=content_hash,
        )

    def insert_batch(self, _target: object, batch: object) -> None:
        self.inserted_row_count += batch.num_rows
        self.batch_rows.append(batch.num_rows)

    def validate_and_promote(self, target: object, **statistics: object) -> str:
        if statistics["row_count"] != self.inserted_row_count:
            raise AssertionError("published row count does not match inserted rows")
        return target.physical_table_name

    def discard_publication(self, _target: object) -> None:
        self.inserted_row_count = 0

    def query(self, statement: str, _parameters: object) -> dict[str, object]:
        self.query_calls += 1
        self.statements.append(statement)
        if self.fail_with is not None:
            raise self.fail_with
        self.returned_row_count = 1
        return dict(self.result)

    def close(self) -> None:
        return None


class CountingClickHouseGateway:
    def __init__(self, delegate: ClickHouseGateway) -> None:
        self.delegate = delegate
        self.query_calls = 0
        self.returned_row_count = 0
        self.statements: list[str] = []

    def __getattr__(self, name: str) -> object:
        return getattr(self.delegate, name)

    def query(self, statement: str, parameters: object) -> object:
        self.query_calls += 1
        self.statements.append(statement)
        result = self.delegate.query(statement, parameters)
        rows = getattr(result, "result_rows", None)
        self.returned_row_count = 1 if rows is None else len(rows)
        return result


def _reference(
    aggregate: str,
    predicate,
) -> dict[str, Decimal | int | None]:
    filtered = [
        _fixture_row(index) for index in range(ROW_COUNT) if predicate(*_fixture_row(index))
    ]
    values = [amount for amount, _region, _order_date in filtered if amount is not None]
    if aggregate == "avg":
        value: Decimal | int | None = sum(values, Decimal(0)) / len(values)
    elif aggregate == "sum":
        value = sum(values, Decimal(0))
    elif aggregate == "count":
        value = len(values)
    elif aggregate == "min":
        value = min(values)
    else:
        value = max(values)
    return {
        "aggregate_value": value,
        "total_count": len(filtered),
        "valid_count": len(values),
        "null_count": len(filtered) - len(values),
    }


class StructuredAggregationEndToEndTest(unittest.TestCase):
    def test_large_summary_matches_decimal_reference_without_python_rows_or_llm(self) -> None:
        expected = _multi_decimal_reference("华东")
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            fixture = _write_multi_fixture(root / "structured-multi-100001.xlsx")
            catalog = sample_multi_metric_catalog()
            schema = replace(
                catalog.datasets[0].schema,
                columns=catalog.datasets[0].schema.columns[:4],
            )
            gateway = LargeMultiSummaryGateway(
                {
                    "total_count": expected["total_count"],
                    "metric_0_value": expected["sales_amount_value"],
                    "metric_0_valid_count": expected["sales_amount_valid_count"],
                    "metric_0_null_count": expected["sales_amount_null_count"],
                    "metric_1_value": expected["cost_amount_value"],
                    "metric_1_valid_count": expected["cost_amount_valid_count"],
                    "metric_1_null_count": expected["cost_amount_null_count"],
                    "metric_2_value": expected["profit_amount_value"],
                    "metric_2_valid_count": expected["profit_amount_valid_count"],
                    "metric_2_null_count": expected["profit_amount_null_count"],
                }
            )
            publication = SpreadsheetPublisher(
                sink=ArrowParquetSink(root / "parquet"),
                clickhouse=gateway,
                batch_rows=25_000,
            ).publish(fixture, schema, "pub-large-multi")
            active_publication = replace(
                catalog.datasets[0].active_publication,
                publication_id=publication.publication_id,
                physical_table_name=publication.physical_table_name,
                row_count=publication.row_count,
                content_hash=publication.content_hash,
            )
            catalog = replace(
                catalog,
                datasets=(
                    replace(
                        catalog.datasets[0],
                        schema=schema,
                        active_publication=active_publication,
                    ),
                ),
            )
            physoc = PhysocFake()
            database = Database("sqlite+pysqlite:///:memory:")
            database.create_schema()
            repository = SqlChatRepository(
                database,
                llm_provider=physoc,
                structured_service=StructuredAnswerService(lambda: catalog, gateway),
            )
            app = create_app(repository=repository, upload_dir=root / "uploads")
            try:
                with TestClient(app) as client:
                    conversation_id = client.post("/api/conversations").json()[
                        "activeConversationId"
                    ]
                    response = client.post(
                        f"/api/conversations/{conversation_id}/messages",
                        json={
                            "content": "地区为华东的销售额、成本、利润汇总",
                            "mode": "source",
                        },
                    )

                self.assertEqual(response.status_code, 200, response.text)
                artifact = response.json()["messages"][-1]["artifacts"][0]
                values = {
                    row[0]: Decimal(row[2].replace(",", ""))
                    for row in artifact["rows"]
                }
                rows_by_metric = {row[0]: row for row in artifact["rows"]}
                self.assertEqual(values["销售额"], expected["sales_amount_value"])
                self.assertEqual(values["成本"], expected["cost_amount_value"])
                self.assertEqual(values["利润"], expected["profit_amount_value"])
                self.assertTrue(
                    all(row[3] == str(expected["total_count"]) for row in artifact["rows"])
                )
                for display_name, field_name in (
                    ("销售额", "sales_amount"),
                    ("成本", "cost_amount"),
                    ("利润", "profit_amount"),
                ):
                    self.assertEqual(
                        rows_by_metric[display_name][4],
                        str(expected[f"{field_name}_valid_count"]),
                    )
                    self.assertEqual(
                        rows_by_metric[display_name][5],
                        str(expected[f"{field_name}_null_count"]),
                    )
                self.assertEqual(gateway.query_calls, 1)
                self.assertEqual(gateway.returned_row_count, 1)
                self.assertEqual(gateway.statements[0].upper().count("SELECT"), 1)
                for alias in ("metric_0_value", "metric_1_value", "metric_2_value"):
                    self.assertIn(alias, gateway.statements[0])
                self.assertEqual(gateway.inserted_row_count, MULTI_ROW_COUNT)
                self.assertNotIn("rows", vars(gateway))
                self.assertEqual(physoc.calls, 0)
            finally:
                database.engine.dispose()

    def test_upload_publish_query_and_fail_closed_routing(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            fixture = _write_fixture(root / "structured-100k.xlsx")
            database = Database("sqlite+pysqlite:///:memory:")
            database.create_schema()
            structured_repository = StructuredRepository(database)
            gateway = InMemoryAggregateGateway()
            catalogs = SwitchableCatalog(structured_repository)
            physoc = PhysocFake()
            chat_repository = SqlChatRepository(
                database,
                llm_provider=physoc,
                structured_service=StructuredAnswerService(catalogs, gateway),
            )
            app = create_app(
                repository=chat_repository,
                structured_repository=structured_repository,
                structured_query_enabled=True,
                upload_dir=root / "uploads",
            )
            try:
                with TestClient(app) as client, fixture.open("rb") as handle:
                    upload = client.post(
                        "/api/knowledge/uploads",
                        files={"files": (fixture.name, handle, XLSX_MIME)},
                    )
                    self.assertEqual(upload.status_code, 200, upload.text)
                    source_id = upload.json()[0]["id"]
                    client.get("/api/knowledge/sources")
                    preview_response = client.get(
                        f"/api/knowledge/sources/{source_id}/structured-preview"
                    )
                    self.assertEqual(preview_response.status_code, 200, preview_response.text)
                    preview = preview_response.json()
                    dataset = preview["datasets"][0]
                    columns = [
                        {
                            "physicalName": column["physicalName"],
                            "displayName": column["physicalName"],
                            "dataType": column["dataType"],
                            "aliases": column["aliases"],
                            "allowAggregate": column["physicalName"] == "order_amount",
                            "allowFilter": True,
                            "nullPolicy": "ignore",
                        }
                        for column in dataset["columns"]
                    ]
                    confirmation = client.put(
                        f"/api/knowledge/sources/{source_id}/structured-schema",
                        json={
                            "datasets": [{"datasetId": dataset["datasetId"], "columns": columns}]
                        },
                    )
                    self.assertEqual(confirmation.status_code, 200, confirmation.text)
                    enqueue = client.post(
                        f"/api/knowledge/sources/{source_id}/structured-publications"
                    )
                    self.assertEqual(enqueue.status_code, 202, enqueue.text)

                    worker = StructuredIngestionWorker(
                        structured_repository,
                        SpreadsheetPublisher(
                            sink=ArrowParquetSink(root / "parquet"),
                            clickhouse=gateway,
                            batch_rows=25_000,
                        ),
                        worker_id="e2e-worker",
                        lease_seconds=60,
                    )
                    self.assertTrue(worker.run_once())
                    status = client.get(f"/api/knowledge/sources/{source_id}/structured-status")
                    self.assertEqual(status.status_code, 200, status.text)
                    self.assertEqual(status.json()["job"]["status"], "published")
                    self.assertEqual(status.json()["activePublication"]["rowCount"], ROW_COUNT)
                    self.assertEqual(sum(gateway.batch_rows), ROW_COUNT)
                    self.assertLessEqual(max(gateway.batch_rows), 25_000)

                    cases = (
                        ("order_amount平均值", "avg", lambda _a, _r, _d: True),
                        ("order_amount总和，region为华东", "sum", lambda _a, r, _d: r == "华东"),
                        ("order_amount计数", "count", lambda _a, _r, _d: True),
                        (
                            "order_amount最小值，order_amount不少于250",
                            "min",
                            lambda amount, _r, _d: amount is not None and amount >= Decimal(250),
                        ),
                        (
                            "order_amount最大值，order_date2025-04-01至2025-06-30",
                            "max",
                            lambda _a, _r, order_date: (
                                date(2025, 4, 1) <= order_date <= date(2025, 6, 30)
                            ),
                        ),
                    )
                    for question, aggregate, predicate in cases:
                        with self.subTest(question=question):
                            conversation_id = client.post("/api/conversations").json()[
                                "activeConversationId"
                            ]
                            response = client.post(
                                f"/api/conversations/{conversation_id}/messages",
                                json={"content": question, "mode": "source"},
                            )
                            self.assertEqual(response.status_code, 200, response.text)
                            expected = _reference(aggregate, predicate)
                            self.assertEqual(gateway.last_result, expected)
                            answer = response.json()["messages"][-1]["paragraphs"][0]["text"]
                            self.assertIn(f"aggregate={aggregate}", answer)
                            self.assertIn(
                                f"value={format(expected['aggregate_value'], ',')}", answer
                            )
                            self.assertIn(f"total={expected['total_count']}", answer)
                            self.assertIn(f"valid={expected['valid_count']}", answer)
                            self.assertIn(f"null={expected['null_count']}", answer)

                    count_all_conversation = client.post("/api/conversations").json()[
                        "activeConversationId"
                    ]
                    count_all = client.post(
                        f"/api/conversations/{count_all_conversation}/messages",
                        json={"content": "总共有多少条记录", "mode": "source"},
                    )
                    self.assertEqual(count_all.status_code, 200, count_all.text)
                    self.assertEqual(
                        gateway.last_result,
                        {
                            "aggregate_value": ROW_COUNT,
                            "total_count": ROW_COUNT,
                            "valid_count": ROW_COUNT,
                            "null_count": 0,
                        },
                    )
                    self.assertIn("count() AS aggregate_value", gateway.query_calls[-1][0])

                    document_conversation = client.post("/api/conversations").json()[
                        "activeConversationId"
                    ]
                    document = client.post(
                        f"/api/conversations/{document_conversation}/messages",
                        json={"content": "请介绍差旅报销政策", "mode": "source"},
                    )
                    self.assertEqual(document.status_code, 200, document.text)
                    self.assertEqual(physoc.calls, 1)
                    self.assertEqual(
                        document.json()["messages"][-1]["paragraphs"][0]["text"],
                        "legacy Physoc answer",
                    )

                    catalogs.override = sample_catalog(ambiguous=True)
                    query_count = len(gateway.query_calls)
                    ambiguous_conversation = client.post("/api/conversations").json()[
                        "activeConversationId"
                    ]
                    ambiguous = client.post(
                        f"/api/conversations/{ambiguous_conversation}/messages",
                        json={"content": "平均金额", "mode": "source"},
                    )
                    self.assertEqual(ambiguous.status_code, 200, ambiguous.text)
                    self.assertEqual(len(gateway.query_calls), query_count)
                    self.assertEqual(physoc.calls, 1)
                    ambiguous_run = next(
                        run
                        for run in chat_repository.list_agent_runs()
                        if run.conversation_id == ambiguous_conversation
                    )
                    self.assertIn(
                        "clarification",
                        ambiguous_run.steps[0].output_summary,
                    )

                    catalogs.override = None
                    gateway.error = TimeoutError("ClickHouse query timed out")
                    timeout_conversation = client.post("/api/conversations").json()[
                        "activeConversationId"
                    ]
                    timeout = client.post(
                        f"/api/conversations/{timeout_conversation}/messages",
                        json={"content": "order_amount总和", "mode": "source"},
                    )
                    self.assertEqual(timeout.status_code, 200, timeout.text)
                    self.assertEqual(physoc.calls, 1)
                    timeout_run = next(
                        run
                        for run in chat_repository.list_agent_runs()
                        if run.conversation_id == timeout_conversation
                    )
                    self.assertIn(
                        "unavailable",
                        timeout_run.steps[0].output_summary,
                    )
            finally:
                database.engine.dispose()


def _create_target_host_clients(
    client_factory,
    ingest_kwargs: dict[str, object],
    query_kwargs: dict[str, object],
):
    ingest = client_factory(**ingest_kwargs)
    try:
        query = client_factory(**query_kwargs)
        if query is ingest:
            raise RuntimeError("target-host gate requires separate ingest and query clients")
    except BaseException as error:
        try:
            ingest.close()
        except BaseException as cleanup_error:
            error.add_note(f"ingest client cleanup also failed: {cleanup_error}")
        raise
    return ingest, query


def _cleanup_target_host_clients(
    ingest: object,
    query: object,
    staging: str,
    table: str,
) -> BaseException | None:
    failures: list[BaseException] = []
    for statement in (f"DROP TABLE IF EXISTS {staging}", f"DROP TABLE IF EXISTS {table}"):
        try:
            ingest.command(statement)
        except BaseException as error:
            failures.append(error)
    for client in (query, ingest):
        try:
            client.close()
        except BaseException as error:
            failures.append(error)
    return failures[0] if failures else None


def _raise_or_note_cleanup(
    primary_error: BaseException | None,
    cleanup_error: BaseException | None,
) -> None:
    if cleanup_error is None:
        return
    if primary_error is None:
        raise cleanup_error
    primary_error.add_note(f"target-host cleanup also failed: {cleanup_error}")


class StructuredAggregationTargetHostHelperTest(unittest.TestCase):
    def test_query_client_creation_failure_closes_ingest_client(self) -> None:
        ingest = type(
            "Ingest",
            (),
            {
                "close_calls": 0,
                "close": lambda self: setattr(self, "close_calls", self.close_calls + 1),
            },
        )()
        calls = 0

        def factory(**_kwargs):
            nonlocal calls
            calls += 1
            if calls == 2:
                raise RuntimeError("query client failed")
            return ingest

        with self.assertRaisesRegex(RuntimeError, "query client failed"):
            _create_target_host_clients(factory, {"role": "ingest"}, {"role": "query"})

        self.assertEqual(ingest.close_calls, 1)

    def test_cleanup_attempts_every_drop_and_close_without_masking_primary_error(self) -> None:
        events = []
        cleanup_failure = RuntimeError("staging drop failed")

        class Ingest:
            def command(self, statement):
                events.append(statement)
                if statement.endswith("staging"):
                    raise cleanup_failure

            def close(self):
                events.append("close-ingest")

        class Query:
            def close(self):
                events.append("close-query")

        cleanup_error = _cleanup_target_host_clients(Ingest(), Query(), "table_staging", "table")

        self.assertIs(cleanup_error, cleanup_failure)
        self.assertEqual(
            events,
            [
                "DROP TABLE IF EXISTS table_staging",
                "DROP TABLE IF EXISTS table",
                "close-query",
                "close-ingest",
            ],
        )
        primary = ValueError("primary query failure")
        _raise_or_note_cleanup(primary, cleanup_error)
        self.assertIn("staging drop failed", primary.__notes__[0])
        with self.assertRaisesRegex(RuntimeError, "staging drop failed"):
            _raise_or_note_cleanup(None, cleanup_error)


def _target_host_missing_configuration() -> tuple[str, ...]:
    required = (
        "RUN_OFFLINE_INTEGRATION",
        "CLICKHOUSE_HOST",
        "CLICKHOUSE_INGEST_PASSWORD_FILE",
        "CLICKHOUSE_QUERY_PASSWORD_FILE",
    )
    missing = []
    for name in required:
        value = os.getenv(name, "").strip()
        if (
            not value
            or (name == "RUN_OFFLINE_INTEGRATION" and value != "1")
            or name.endswith("_FILE")
            and not Path(value).is_file()
        ):
            missing.append(name)
    return tuple(missing)


_TARGET_HOST_MISSING = _target_host_missing_configuration()


@unittest.skipUnless(
    not _TARGET_HOST_MISSING,
    "target-host gate missing explicit configuration: " + ", ".join(_TARGET_HOST_MISSING),
)
class StructuredAggregationTargetHostGateTest(unittest.TestCase):
    def test_large_filtered_multi_summary_matches_decimal_reference(self) -> None:
        import clickhouse_connect

        expected = _multi_decimal_reference("华东")
        host = os.environ["CLICKHOUSE_HOST"]
        port = int(os.getenv("CLICKHOUSE_PORT", "8123"))
        ingest, query = _create_target_host_clients(
            clickhouse_connect.get_client,
            {
                "host": host,
                "port": port,
                "username": os.getenv("CLICKHOUSE_INGEST_USER", "structured_ingest"),
                "password": require_secret_file(
                    Path(os.environ["CLICKHOUSE_INGEST_PASSWORD_FILE"]),
                    "CLICKHOUSE_INGEST_PASSWORD_FILE",
                ),
            },
            {
                "host": host,
                "port": port,
                "username": os.getenv("CLICKHOUSE_QUERY_USER", "structured_query"),
                "password": require_secret_file(
                    Path(os.environ["CLICKHOUSE_QUERY_PASSWORD_FILE"]),
                    "CLICKHOUSE_QUERY_PASSWORD_FILE",
                ),
                "autogenerate_session_id": False,
            },
        )
        primary_error: BaseException | None = None
        published_table = ""
        database: Database | None = None
        try:
            with tempfile.TemporaryDirectory() as directory:
                root = Path(directory)
                fixture = _write_multi_fixture(root / "target-multi-100001.xlsx")
                catalog = sample_multi_metric_catalog()
                schema = replace(
                    catalog.datasets[0].schema,
                    dataset_id=f"ds_target_{uuid.uuid4().hex}",
                    source_id=f"kb_target_{uuid.uuid4().hex}",
                    columns=catalog.datasets[0].schema.columns[:4],
                )
                gateway = CountingClickHouseGateway(
                    ClickHouseGateway(ingest, query_client=query, max_result_rows=1)
                )
                publication = SpreadsheetPublisher(
                    sink=ArrowParquetSink(root / "parquet"),
                    clickhouse=gateway,
                    batch_rows=25_000,
                ).publish(fixture, schema, f"pub-target-{uuid.uuid4().hex}")
                published_table = publication.physical_table_name
                active_publication = replace(
                    catalog.datasets[0].active_publication,
                    publication_id=publication.publication_id,
                    dataset_id=schema.dataset_id,
                    physical_table_name=publication.physical_table_name,
                    row_count=publication.row_count,
                    content_hash=publication.content_hash,
                )
                catalog = replace(
                    catalog,
                    datasets=(
                        replace(
                            catalog.datasets[0],
                            schema=schema,
                            active_publication=active_publication,
                        ),
                    ),
                )
                physoc = PhysocFake()
                database = Database("sqlite+pysqlite:///:memory:")
                database.create_schema()
                repository = SqlChatRepository(
                    database,
                    llm_provider=physoc,
                    structured_service=StructuredAnswerService(lambda: catalog, gateway),
                )
                app = create_app(repository=repository, upload_dir=root / "uploads")
                with TestClient(app) as client:
                    conversation_id = client.post("/api/conversations").json()[
                        "activeConversationId"
                    ]
                    response = client.post(
                        f"/api/conversations/{conversation_id}/messages",
                        json={
                            "content": "地区为华东的销售额、成本、利润汇总",
                            "mode": "source",
                        },
                    )

                self.assertEqual(response.status_code, 200, response.text)
                artifact = response.json()["messages"][-1]["artifacts"][0]
                rows_by_metric = {row[0]: row for row in artifact["rows"]}
                for display_name, field_name in (
                    ("销售额", "sales_amount"),
                    ("成本", "cost_amount"),
                    ("利润", "profit_amount"),
                ):
                    row = rows_by_metric[display_name]
                    self.assertEqual(
                        Decimal(row[2].replace(",", "")),
                        expected[f"{field_name}_value"],
                    )
                    self.assertEqual(row[3], str(expected["total_count"]))
                    self.assertEqual(row[4], str(expected[f"{field_name}_valid_count"]))
                    self.assertEqual(row[5], str(expected[f"{field_name}_null_count"]))
                self.assertEqual(gateway.query_calls, 1)
                self.assertEqual(gateway.returned_row_count, 1)
                self.assertEqual(gateway.statements[0].upper().count("SELECT"), 1)
                self.assertEqual(physoc.calls, 0)
                stored_count = query.query(
                    f"SELECT count() FROM {published_table}"
                ).result_rows[0][0]
                self.assertEqual(stored_count, MULTI_ROW_COUNT)
        except BaseException as error:
            primary_error = error
            raise
        finally:
            if database is not None:
                database.engine.dispose()
            if published_table:
                cleanup_error = _cleanup_target_host_clients(
                    ingest,
                    query,
                    f"{published_table}_cleanup_staging",
                    published_table,
                )
            else:
                cleanup_error = None
                for client in (query, ingest):
                    try:
                        client.close()
                    except BaseException as error:
                        cleanup_error = cleanup_error or error
            _raise_or_note_cleanup(primary_error, cleanup_error)

    def test_clickhouse_target_publishes_and_queries_with_separate_identities(self) -> None:
        import clickhouse_connect

        table = f"structured_e2e_gate_{uuid.uuid4().hex[:16]}"
        staging = f"{table}_staging"
        host = os.environ["CLICKHOUSE_HOST"]
        port = int(os.getenv("CLICKHOUSE_PORT", "8123"))
        ingest, query = _create_target_host_clients(
            clickhouse_connect.get_client,
            {
                "host": host,
                "port": port,
                "username": os.getenv("CLICKHOUSE_INGEST_USER", "structured_ingest"),
                "password": require_secret_file(
                    Path(os.environ["CLICKHOUSE_INGEST_PASSWORD_FILE"]),
                    "CLICKHOUSE_INGEST_PASSWORD_FILE",
                ),
            },
            {
                "host": host,
                "port": port,
                "username": os.getenv("CLICKHOUSE_QUERY_USER", "structured_query"),
                "password": require_secret_file(
                    Path(os.environ["CLICKHOUSE_QUERY_PASSWORD_FILE"]),
                    "CLICKHOUSE_QUERY_PASSWORD_FILE",
                ),
                "autogenerate_session_id": False,
            },
        )
        primary_error: BaseException | None = None
        try:
            ingest.command(f"DROP TABLE IF EXISTS {staging}")
            ingest.command(f"DROP TABLE IF EXISTS {table}")
            ingest.command(
                f"CREATE TABLE {staging} (amount Nullable(Decimal(38, 9))) "
                "ENGINE = MergeTree ORDER BY tuple()"
            )
            ingest.insert(
                staging,
                [[Decimal("10")], [None], [Decimal("30")]],
                column_names=("amount",),
            )
            ingest.command(f"RENAME TABLE {staging} TO {table}")
            result = query.query(f"SELECT count(), avg(amount) FROM {table}").result_rows[0]
            self.assertEqual(result, (3, Decimal("20")))
        except BaseException as error:
            primary_error = error
            raise
        finally:
            _raise_or_note_cleanup(
                primary_error,
                _cleanup_target_host_clients(ingest, query, staging, table),
            )


if __name__ == "__main__":
    unittest.main()
