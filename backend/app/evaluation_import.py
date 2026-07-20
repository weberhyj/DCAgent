from __future__ import annotations

import csv
import io
import json
import re
import secrets
import threading
import time
from collections.abc import Iterable
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .evaluation import normalize_evaluation_case_metadata

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_FILE_NAME_LENGTH = 240
MAX_IMPORT_ROWS = 2000
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_STORED_PREVIEWS = 100


class EvaluationImportFileError(ValueError):
    pass


class EvaluationImportTokenError(ValueError):
    pass


class EvaluationImportTokenBusyError(EvaluationImportTokenError):
    pass


@dataclass(slots=True)
class EvaluationImportRow:
    row_number: int
    question: str
    expect_answer: bool
    expected_source_ids: list[str]
    expected_terms: list[str]
    category: str | None
    tags: list[str]
    top_k: int
    external_key: str | None


@dataclass(slots=True)
class EvaluationImportError:
    row_number: int
    field: str
    message: str


@dataclass(slots=True)
class EvaluationImportPreview:
    token: str
    file_name: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_rows: int
    rows: list[EvaluationImportRow]
    errors: list[EvaluationImportError]
    duplicate_keys: list[str]
    expires_at: float


def parse_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        if value == 1:
            return True
        if value == 0:
            return False

    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "是", "应有答案"}:
        return True
    if normalized in {"false", "0", "否", "应无答案"}:
        return False
    raise ValueError("答案预期仅支持 true/false、1/0、是/否、应有答案/应无答案")


def split_values(value: object) -> list[str]:
    if value is None:
        return []
    values: Iterable[object]
    if isinstance(value, list):
        values = value
    elif isinstance(value, str):
        values = value.split("|")
    else:
        raise ValueError("值必须是列表或字符串")
    if any(isinstance(item, (dict, list, tuple, set)) for item in values):
        raise ValueError("列表项必须是标量")
    return list(dict.fromkeys(str(item).strip() for item in values if str(item).strip()))


def read_import_records(file_name: str, content: bytes) -> list[dict[str, object]]:
    if len(file_name) > MAX_IMPORT_FILE_NAME_LENGTH:
        raise EvaluationImportFileError("文件名不能超过 240 个字符")
    if not content:
        raise EvaluationImportFileError("文件不能为空")
    if len(content) > MAX_IMPORT_BYTES:
        raise EvaluationImportFileError("文件不能超过 5MB")

    suffix = Path(file_name).suffix.lower()
    try:
        if suffix == ".csv":
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            records = _read_bounded_records(reader)
        elif suffix == ".json":
            payload = json.loads(content.decode("utf-8-sig"))
            if not isinstance(payload, list):
                raise EvaluationImportFileError("JSON 顶层必须是数组")
            if len(payload) > MAX_IMPORT_ROWS:
                raise EvaluationImportFileError("导入数据不能超过 2000 行")
            records = payload
        elif suffix == ".xlsx":
            _validate_xlsx_archive(content)
            records = _read_xlsx_records(content)
        else:
            raise EvaluationImportFileError("仅支持 XLSX、CSV 和 JSON 文件")
    except EvaluationImportFileError:
        raise
    except UnicodeDecodeError as exc:
        raise EvaluationImportFileError("文件必须使用 UTF-8 编码") from exc
    except json.JSONDecodeError as exc:
        raise EvaluationImportFileError("JSON 文件格式无效") from exc
    except csv.Error as exc:
        raise EvaluationImportFileError("CSV 文件格式无效") from exc
    except (BadZipFile, InvalidFileException) as exc:
        raise EvaluationImportFileError("XLSX 文件格式无效") from exc
    except Exception as exc:
        if suffix == ".xlsx":
            raise EvaluationImportFileError("XLSX 文件格式无效") from exc
        raise

    if not records:
        raise EvaluationImportFileError("文件中无有效数据")
    return records


def _read_bounded_records(records: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    bounded = list(islice(records, MAX_IMPORT_ROWS + 1))
    if len(bounded) > MAX_IMPORT_ROWS:
        raise EvaluationImportFileError("导入数据不能超过 2000 行")
    return bounded


def _validate_xlsx_archive(content: bytes) -> None:
    with ZipFile(io.BytesIO(content)) as archive:
        total_uncompressed_bytes = sum(entry.file_size for entry in archive.infolist())
    if total_uncompressed_bytes > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise EvaluationImportFileError("XLSX 文件解压后不能超过 50MB")


def _read_xlsx_records(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers_row = next(rows, None)
        if headers_row is None:
            return []
        headers = [str(value or "").strip() for value in headers_row]
        return _read_bounded_records(dict(zip(headers, values, strict=False)) for values in rows)
    finally:
        workbook.close()


def _record_value(record: dict[str, object], field: str, *aliases: str) -> object:
    normalized_record = {
        str(key).strip(): value for key, value in record.items() if key is not None
    }
    for name in (field, *aliases):
        if name in normalized_record:
            return normalized_record[name]
    return None


def _optional_text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _require_scalar(value: object, field: str, label: str) -> None:
    if isinstance(value, (dict, list, tuple, set)):
        raise _RowValueError(field, f"{label}必须是标量值")


def _parse_values_field(
    value: object,
    field: str,
    label: str,
) -> list[str]:
    try:
        return split_values(value)
    except ValueError as exc:
        raise _RowValueError(field, f"{label}必须是列表或字符串") from exc


def _clean_question(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalize_question_key(value: object) -> str:
    return _clean_question(value).casefold()


def _parse_top_k(value: object) -> int:
    if value is None or (isinstance(value, str) and not value.strip()):
        return 5
    if isinstance(value, bool):
        raise ValueError("top_k 必须是 1 到 10 的整数")
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("top_k 必须是 1 到 10 的整数") from exc
    if isinstance(value, float) and not value.is_integer():
        raise ValueError("top_k 必须是 1 到 10 的整数")
    if not 1 <= parsed <= 10:
        raise ValueError("top_k 必须是 1 到 10 的整数")
    return parsed


class EvaluationImportService:
    def __init__(
        self,
        ttl_seconds: int = 1800,
        max_previews: int = MAX_STORED_PREVIEWS,
    ) -> None:
        self._ttl_seconds = ttl_seconds
        self._max_previews = max_previews
        self._previews: dict[str, EvaluationImportPreview] = {}
        self._reserved_tokens: set[str] = set()
        self._lock = threading.Lock()

    @property
    def ttl_seconds(self) -> int:
        return self._ttl_seconds

    def preview(
        self,
        file_name: str,
        content: bytes,
        sources: Iterable[object],
        existing_cases: Iterable[object] | None = None,
    ) -> EvaluationImportPreview:
        with self._lock:
            self._purge_expired_locked(time.time())
            self._require_capacity_locked()

        records = read_import_records(file_name, content)
        existing_case_snapshot = tuple(existing_cases or ())
        source_ids_by_name: dict[str, list[str]] = {}
        for source in sources:
            source_name = str(getattr(source, "name", "")).strip()
            source_id = str(getattr(source, "id", "")).strip()
            if source_name and source_id:
                source_ids_by_name.setdefault(source_name, []).append(source_id)

        existing_external_keys = {
            str(getattr(case, "external_key", "") or "").strip()
            for case in existing_case_snapshot
            if str(getattr(case, "external_key", "") or "").strip()
        }
        existing_questions = {
            _normalize_question_key(getattr(case, "question", ""))
            for case in existing_case_snapshot
            if _normalize_question_key(getattr(case, "question", ""))
        }

        rows: list[EvaluationImportRow] = []
        errors: list[EvaluationImportError] = []
        duplicate_keys: list[str] = []
        for row_number, raw_record in enumerate(records, start=2):
            if not isinstance(raw_record, dict):
                errors.append(EvaluationImportError(row_number, "row", "每行数据必须是字段对象"))
                continue
            try:
                row = self._parse_row(row_number, raw_record, source_ids_by_name)
            except ValueError as exc:
                field = getattr(exc, "field", "row")
                errors.append(EvaluationImportError(row_number, field, str(exc)))
                continue

            rows.append(row)
            duplicate_key = None
            if row.external_key:
                if row.external_key in existing_external_keys:
                    duplicate_key = row.external_key
            else:
                normalized_question = _normalize_question_key(row.question)
                if normalized_question in existing_questions:
                    duplicate_key = normalized_question
            if duplicate_key and duplicate_key not in duplicate_keys:
                duplicate_keys.append(duplicate_key)

        with self._lock:
            now = time.time()
            self._purge_expired_locked(now)
            self._require_capacity_locked()
            token = secrets.token_urlsafe(24)
            while token in self._previews:
                token = secrets.token_urlsafe(24)
            preview = EvaluationImportPreview(
                token=token,
                file_name=file_name,
                total_rows=len(records),
                valid_rows=len(rows),
                invalid_rows=len(records) - len(rows),
                duplicate_rows=sum(
                    1
                    for row in rows
                    if (
                        row.external_key in existing_external_keys
                        if row.external_key
                        else _normalize_question_key(row.question) in existing_questions
                    )
                ),
                rows=rows,
                errors=errors,
                duplicate_keys=duplicate_keys,
                expires_at=now + self._ttl_seconds,
            )
            self._previews[token] = preview
        return preview

    def reserve(self, token: str) -> EvaluationImportPreview:
        with self._lock:
            self._purge_expired_locked(time.time())
            preview = self._previews.get(token)
            if preview is not None and token in self._reserved_tokens:
                raise EvaluationImportTokenBusyError("导入预览正在确认，请稍后重试")
            if preview is not None:
                self._reserved_tokens.add(token)
        if preview is None:
            raise EvaluationImportTokenError("导入预览已过期，请重新上传文件")
        return preview

    def complete(self, token: str) -> None:
        with self._lock:
            self._reserved_tokens.discard(token)
            self._previews.pop(token, None)

    def release(self, token: str) -> None:
        with self._lock:
            self._reserved_tokens.discard(token)
            preview = self._previews.get(token)
            if preview is not None and preview.expires_at <= time.time():
                self._previews.pop(token, None)

    def consume(self, token: str) -> EvaluationImportPreview:
        preview = self.reserve(token)
        self.complete(token)
        return preview

    def _purge_expired_locked(self, now: float) -> None:
        expired_tokens = [
            token for token, preview in self._previews.items() if preview.expires_at <= now
        ]
        for token in expired_tokens:
            self._previews.pop(token, None)
            self._reserved_tokens.discard(token)

    def _require_capacity_locked(self) -> None:
        if len(self._previews) >= self._max_previews:
            raise EvaluationImportFileError("导入预览数量已达上限，请稍后重试")

    def _parse_row(
        self,
        row_number: int,
        record: dict[str, object],
        source_ids_by_name: dict[str, list[str]],
    ) -> EvaluationImportRow:
        question_value = _record_value(record, "question")
        _require_scalar(question_value, "question", "问题")
        question = _clean_question(question_value)
        if not question:
            raise _RowValueError("question", "问题不能为空")
        if len(question) > 1000:
            raise _RowValueError("question", "问题不能超过 1000 个字符")

        expect_answer_value = _record_value(
            record,
            "expect_answer",
            "expectAnswer",
        )
        if expect_answer_value is None or (
            isinstance(expect_answer_value, str) and not expect_answer_value.strip()
        ):
            raise _RowValueError("expect_answer", "答案预期必须显式提供且不能为空")
        _require_scalar(expect_answer_value, "expect_answer", "答案预期")
        try:
            expect_answer = parse_boolean(expect_answer_value)
        except ValueError as exc:
            raise _RowValueError("expect_answer", str(exc)) from exc

        source_names = _parse_values_field(
            _record_value(record, "expected_sources", "expectedSources"),
            "expected_sources",
            "期望资料",
        )
        expected_source_ids: list[str] = []
        for source_name in source_names:
            source_ids = source_ids_by_name.get(source_name, [])
            if not source_ids:
                raise _RowValueError(
                    "expected_sources",
                    f"未找到资料：{source_name}",
                )
            if len(source_ids) > 1:
                raise _RowValueError(
                    "expected_sources",
                    f"存在多个同名资料：{source_name}",
                )
            if source_ids[0] not in expected_source_ids:
                expected_source_ids.append(source_ids[0])

        expected_terms = _parse_values_field(
            _record_value(record, "expected_terms", "expectedTerms"),
            "expected_terms",
            "期望关键词",
        )
        if expect_answer and not expected_source_ids and not expected_terms:
            raise _RowValueError(
                "expected_evidence",
                "应有答案的问题必须提供期望资料或关键词",
            )

        top_k_value = _record_value(record, "top_k", "topK")
        _require_scalar(top_k_value, "top_k", "top_k")
        try:
            top_k = _parse_top_k(top_k_value)
        except ValueError as exc:
            raise _RowValueError("top_k", str(exc)) from exc

        raw_category = _record_value(record, "category")
        _require_scalar(raw_category, "category", "分类")
        category_value = _optional_text(raw_category)
        tags_value = _parse_values_field(
            _record_value(record, "tags"),
            "tags",
            "标签",
        )
        raw_external_key = _record_value(record, "external_key", "externalKey")
        _require_scalar(raw_external_key, "external_key", "外部标识")
        external_key_value = _optional_text(raw_external_key)
        try:
            category, tags, external_key, _ = normalize_evaluation_case_metadata(
                category=category_value,
                tags=tags_value,
                external_key=external_key_value,
            )
        except ValueError as exc:
            raise _metadata_row_error(exc) from exc

        return EvaluationImportRow(
            row_number=row_number,
            question=question,
            expect_answer=expect_answer,
            expected_source_ids=expected_source_ids,
            expected_terms=expected_terms,
            category=category,
            tags=tags,
            top_k=top_k,
            external_key=external_key,
        )


class _RowValueError(ValueError):
    def __init__(self, field: str, message: str) -> None:
        super().__init__(message)
        self.field = field


def _metadata_row_error(error: ValueError) -> _RowValueError:
    message = str(error)
    if message.startswith("category"):
        return _RowValueError("category", "分类不能超过 80 个字符")
    if message.startswith("external_key"):
        return _RowValueError("external_key", "外部标识不能超过 120 个字符")
    if "at most 20 items" in message:
        return _RowValueError("tags", "标签最多 20 个")
    if "at most 80 characters" in message:
        return _RowValueError("tags", "单个标签不能超过 80 个字符")
    return _RowValueError("metadata", "评测元数据格式无效")
