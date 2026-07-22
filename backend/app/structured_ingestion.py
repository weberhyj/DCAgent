from __future__ import annotations

import codecs
import csv
import hashlib
import re
import shutil
from collections.abc import Callable, Iterable, Iterator, Mapping, Sequence
from datetime import date, datetime
from decimal import ROUND_HALF_EVEN, Decimal, InvalidOperation, localcontext
from pathlib import Path
from typing import Any, Protocol

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from .clickhouse_gateway import _content_hash_from_observation, _row_digest_lanes
from .spreadsheet_schema import (
    ENCODING_PROBE_BYTES,
    _bounded_csv_field_size,
    _BoundedCsvLines,
    _candidate_csv_encodings,
    _CsvRecordLimitExceeded,
    _normalize_physical_name,
)
from .structured_models import (
    StructuredColumnSchema,
    StructuredColumnType,
    StructuredDatasetSchema,
    StructuredPublicationResult,
)

MAX_BATCH_ROWS = 50_000
MAX_BATCH_BYTES = 64 * 1024 * 1024
_SAFE_PATH_COMPONENT = re.compile(r"[^A-Za-z0-9._-]+")


class StructuredIngestionError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        row_number: int | None = None,
        physical_column: str | None = None,
        sample: Any = None,
    ) -> None:
        self.code = code
        self.row_number = row_number
        self.physical_column = physical_column
        self.sample = _redacted_sample(sample) if sample is not None else None
        details = [message]
        if row_number is not None:
            details.append(f"row={row_number}")
        if physical_column:
            details.append(f"column={physical_column}")
        if self.sample is not None:
            details.append(f"sample={self.sample}")
        super().__init__("; ".join(details))


class ParquetBatchSink(Protocol):
    root: Path

    def write_batch(self, batch: pa.RecordBatch, output_path: Path) -> None: ...

    def iter_batches(self, paths: Sequence[Path]) -> Iterable[pa.RecordBatch]: ...


class ArrowParquetSink:
    def __init__(self, root: Path) -> None:
        self.root = Path(root)

    def write_batch(self, batch: pa.RecordBatch, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        pq.write_table(pa.Table.from_batches([batch]), output_path, compression="zstd")

    def iter_batches(self, paths: Sequence[Path]) -> Iterable[pa.RecordBatch]:
        for path in paths:
            parquet_file = pq.ParquetFile(path)
            yield from parquet_file.iter_batches()


class SpreadsheetPublisher:
    def __init__(
        self,
        *,
        sink: ParquetBatchSink | None = None,
        clickhouse: Any,
        parquet_root: Path | None = None,
        batch_rows: int = 50_000,
        batch_bytes: int = MAX_BATCH_BYTES,
    ) -> None:
        if not 1 <= batch_rows <= MAX_BATCH_ROWS:
            raise ValueError(f"batch_rows must be between 1 and {MAX_BATCH_ROWS}")
        if not 1 <= batch_bytes <= MAX_BATCH_BYTES:
            raise ValueError(f"batch_bytes must be between 1 and {MAX_BATCH_BYTES}")
        self.sink = sink or ArrowParquetSink(parquet_root or Path("./data/parquet"))
        self.clickhouse = clickhouse
        self.batch_rows = batch_rows
        self.batch_bytes = batch_bytes

    def publish(
        self,
        path: Path,
        schema: StructuredDatasetSchema,
        publication_id: str,
        *,
        lease_guard: Callable[[], None] | None = None,
    ) -> StructuredPublicationResult:
        _check_lease(lease_guard)
        source_root = Path(getattr(self.sink, "root", Path("./data/parquet")))
        output_root = (
            source_root
            / _safe_path_component(schema.source_id)
            / _safe_path_component(schema.dataset_id)
            / str(schema.schema_version)
            / _safe_path_component(publication_id)
        )
        _require_within_root(source_root, output_root)
        _clear_publication_directory(output_root)
        output_paths: list[Path] = []
        buffered: list[dict[str, Any]] = []
        buffered_bytes = 0
        null_counts = {column.physical_name: 0 for column in schema.columns}
        numeric_ranges: dict[str, list[Decimal | int | None]] = {
            column.physical_name: [None, None]
            for column in schema.columns
            if column.data_type in {StructuredColumnType.INTEGER, StructuredColumnType.DECIMAL}
        }
        content_sums = [0, 0, 0, 0]
        content_xors = [0, 0, 0, 0]
        row_count = 0
        column_schema = _arrow_schema(schema.columns)

        try:
            for row_number, values, formulas in _iter_source_rows(Path(path), schema):
                _check_lease(lease_guard)
                if _row_is_empty(values, formulas):
                    continue
                row: dict[str, Any] = {}
                for index, column in enumerate(schema.columns):
                    cached = values[index] if index < len(values) else None
                    formula = formulas[index] if index < len(formulas) else None
                    converted = _convert_value(
                        cached,
                        formula,
                        column,
                        row_number=row_number,
                    )
                    row[column.physical_name] = converted
                    if converted is None:
                        null_counts[column.physical_name] += 1
                    if (
                        column.data_type
                        in {
                            StructuredColumnType.INTEGER,
                            StructuredColumnType.DECIMAL,
                        }
                        and converted is not None
                    ):
                        bounds = numeric_ranges[column.physical_name]
                        bounds[0] = converted if bounds[0] is None else min(bounds[0], converted)
                        bounds[1] = converted if bounds[1] is None else max(bounds[1], converted)

                row.update(
                    {
                        "_source_id": schema.source_id,
                        "_dataset_id": schema.dataset_id,
                        "_schema_version": schema.schema_version,
                        "_worksheet": schema.worksheet_name,
                        "_row_number": row_number,
                    }
                )
                canonical_row = _canonical_row(row)
                if len(canonical_row) > self.batch_bytes:
                    raise StructuredIngestionError(
                        "row_size_limit_exceeded",
                        "Converted row exceeds the configured ingestion batch byte limit.",
                        row_number=row_number,
                    )
                for index, lane in enumerate(_row_digest_lanes(canonical_row)):
                    content_sums[index] = (content_sums[index] + lane) & ((1 << 64) - 1)
                    content_xors[index] ^= lane
                if buffered and (
                    len(buffered) >= self.batch_rows
                    or buffered_bytes + len(canonical_row) > self.batch_bytes
                ):
                    _check_lease(lease_guard)
                    output_paths.append(
                        self._write_batch(buffered, column_schema, output_root, len(output_paths))
                    )
                    buffered = []
                    buffered_bytes = 0
                buffered.append(row)
                buffered_bytes += len(canonical_row)
                row_count += 1
                if len(buffered) >= self.batch_rows or buffered_bytes >= self.batch_bytes:
                    _check_lease(lease_guard)
                    output_paths.append(
                        self._write_batch(buffered, column_schema, output_root, len(output_paths))
                    )
                    buffered = []
                    buffered_bytes = 0
            if buffered:
                _check_lease(lease_guard)
                output_paths.append(
                    self._write_batch(buffered, column_schema, output_root, len(output_paths))
                )
        except Exception:
            _clear_publication_directory(output_root, best_effort=True)
            raise

        content_hash = _content_hash_from_observation(
            row_count=row_count,
            sums=content_sums,
            xors=content_xors,
        )
        target = None
        try:
            _check_lease(lease_guard)
            target = self.clickhouse.prepare_publication(schema, publication_id, content_hash)
            for batch in self.sink.iter_batches(output_paths):
                _check_lease(lease_guard)
                batches = list(batch.columns)
                batches.append(pa.array([content_hash] * batch.num_rows, type=pa.string()))
                batch_schema = pa.schema(
                    list(batch.schema) + [pa.field("_content_hash", pa.string())]
                )
                self.clickhouse.insert_batch(
                    target,
                    pa.RecordBatch.from_arrays(batches, schema=batch_schema),
                )

            numeric_range_values = {
                name: (bounds[0], bounds[1]) for name, bounds in numeric_ranges.items()
            }
            _check_lease(lease_guard)
            physical_table_name = self.clickhouse.validate_and_promote(
                target,
                row_count=row_count,
                column_count=len(schema.columns),
                null_counts=null_counts,
                numeric_ranges=numeric_range_values,
                content_hash=content_hash,
            )
            _check_lease(lease_guard)
        except Exception:
            discard = getattr(self.clickhouse, "discard_publication", None)
            if discard is not None and target is not None:
                try:
                    discard(target)
                except Exception:
                    pass
            _clear_publication_directory(output_root, best_effort=True)
            raise
        return StructuredPublicationResult(
            publication_id=publication_id,
            physical_table_name=physical_table_name,
            row_count=row_count,
            column_count=len(schema.columns),
            null_counts=null_counts,
            content_hash=content_hash,
        )

    def _write_batch(
        self,
        rows: Sequence[dict[str, Any]],
        schema: pa.Schema,
        output_root: Path,
        part_number: int,
    ) -> Path:
        batch = pa.RecordBatch.from_pylist(rows, schema=schema)
        output_path = output_root / f"part-{part_number:05d}.parquet"
        self.sink.write_batch(batch, output_path)
        return output_path


def _check_lease(lease_guard: Callable[[], None] | None) -> None:
    if lease_guard is not None:
        lease_guard()


def _iter_source_rows(
    path: Path,
    schema: StructuredDatasetSchema,
) -> Iterator[tuple[int, tuple[Any, ...], tuple[Any, ...]]]:
    if path.suffix.lower() == ".csv":
        encoding = _validated_csv_encoding(path)
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                header_seen = False
                column_indexes: tuple[int, ...] = ()
                with _bounded_csv_field_size():
                    reader = csv.reader(_BoundedCsvLines(handle, encoding))
                    for row_number, row in enumerate(reader, 1):
                        values = tuple(row)
                        if not header_seen:
                            if _row_is_empty(values, ()):
                                continue
                            column_indexes = _confirmed_column_indexes(values, (), schema)
                            header_seen = True
                            continue
                        yield row_number, _select_columns(values, column_indexes), ()
        except _CsvRecordLimitExceeded as error:
            raise StructuredIngestionError(
                "csv_record_limit_exceeded", "CSV record exceeded the configured byte limit."
            ) from error
        except csv.Error as error:
            raise StructuredIngestionError(
                "csv_read_error", f"CSV could not be read: {error}"
            ) from error
        return
    if path.suffix.lower() != ".xlsx":
        raise StructuredIngestionError(
            "unsupported_format", f"Unsupported spreadsheet format: {path.suffix}"
        )

    cached = load_workbook(path, read_only=True, data_only=True)
    formulas = load_workbook(path, read_only=True, data_only=False)
    cached_rows = formula_rows = None
    try:
        cached_sheet = cached[schema.worksheet_name]
        formula_sheet = formulas[schema.worksheet_name]
        cached_rows = (tuple(cell.value for cell in row) for row in cached_sheet.iter_rows())
        formula_rows = (tuple(cell.value for cell in row) for row in formula_sheet.iter_rows())
        header_seen = False
        column_indexes: tuple[int, ...] = ()
        for row_number, (values, formula_values) in enumerate(
            zip(cached_rows, formula_rows, strict=False), 1
        ):
            if not header_seen:
                if _row_is_empty(values, formula_values):
                    continue
                column_indexes = _confirmed_column_indexes(values, formula_values, schema)
                header_seen = True
                continue
            yield (
                row_number,
                _select_columns(values, column_indexes),
                _select_columns(formula_values, column_indexes),
            )
    finally:
        if cached_rows is not None:
            cached_rows.close()
        if formula_rows is not None:
            formula_rows.close()
        formulas.close()
        cached.close()


def _validated_csv_encoding(path: Path) -> str:
    for encoding in _candidate_csv_encodings(path):
        decoder = codecs.getincrementaldecoder(encoding)(errors="strict")
        try:
            with path.open("rb") as handle:
                while chunk := handle.read(ENCODING_PROBE_BYTES):
                    decoder.decode(chunk, final=False)
            decoder.decode(b"", final=True)
        except UnicodeDecodeError:
            continue
        return encoding
    raise StructuredIngestionError(
        "unsupported_encoding", "CSV could not be decoded with a supported encoding."
    )


def _confirmed_column_indexes(
    values: Sequence[Any],
    formulas: Sequence[Any],
    schema: StructuredDatasetSchema,
) -> tuple[int, ...]:
    width = max(len(values), len(formulas))
    while width > 0:
        cached = values[width - 1] if width <= len(values) else None
        formula = formulas[width - 1] if width <= len(formulas) else None
        if cached is not None or formula is not None:
            break
        width -= 1

    used_names: set[str] = set()
    source_columns: list[tuple[str, str, int]] = []
    for index in range(width):
        cached = values[index] if index < len(values) else None
        formula = formulas[index] if index < len(formulas) else None
        raw = cached
        if (
            raw is None
            and formula is not None
            and not (isinstance(formula, str) and formula.startswith("="))
        ):
            raw = formula
        original = "" if raw is None else str(raw).strip()
        base = _normalize_physical_name(original, index + 1)
        physical_name = base
        suffix = 2
        while physical_name in used_names:
            physical_name = f"{base}_{suffix}"
            suffix += 1
        used_names.add(physical_name)
        source_columns.append((physical_name, original, index))

    selected: list[int] = []
    assigned: set[int] = set()
    missing: list[str] = []
    for column in schema.columns:
        physical_matches = [
            index
            for physical_name, _, index in source_columns
            if physical_name == column.physical_name and index not in assigned
        ]
        original_matches = [
            index
            for _, original_name, index in source_columns
            if original_name == column.original_name and index not in assigned
        ]
        matches = physical_matches or original_matches
        if len(matches) != 1:
            missing.append(column.physical_name)
            continue
        selected.append(matches[0])
        assigned.add(matches[0])

    unexpected = [
        physical_name for physical_name, _, index in source_columns if index not in assigned
    ]
    if missing or unexpected:
        raise StructuredIngestionError(
            "schema_drift",
            f"Source headers changed after confirmation; missing={missing}, unexpected={unexpected}.",
        )
    return tuple(selected)


def _select_columns(values: Sequence[Any], indexes: Sequence[int]) -> tuple[Any, ...]:
    return tuple(values[index] if index < len(values) else None for index in indexes)


def _convert_value(
    value: Any,
    formula: Any,
    column: StructuredColumnSchema,
    *,
    row_number: int,
) -> Any:
    if isinstance(formula, str) and formula.startswith("=") and value is None:
        if column.allow_aggregate:
            raise StructuredIngestionError(
                "formula_cache_missing",
                "Formula aggregate has no cached value; recalculate and save the workbook in an office application.",
                row_number=row_number,
                physical_column=column.physical_name,
            )
        value = None
    if value is None or (isinstance(value, str) and not value.strip()):
        if column.null_policy == "reject":
            raise StructuredIngestionError(
                "null_value_rejected",
                "Null value is not allowed by the confirmed schema.",
                row_number=row_number,
                physical_column=column.physical_name,
            )
        if column.null_policy == "zero" and column.data_type in {
            StructuredColumnType.INTEGER,
            StructuredColumnType.DECIMAL,
        }:
            return 0 if column.data_type is StructuredColumnType.INTEGER else _decimal_38_9("0")
        return None
    try:
        if column.data_type is StructuredColumnType.STRING:
            return str(value)
        if column.data_type is StructuredColumnType.INTEGER:
            integer = int(value)
            if isinstance(value, float) and not value.is_integer():
                raise ValueError("fractional value")
            return integer
        if column.data_type is StructuredColumnType.DECIMAL:
            return _decimal_38_9(value)
        if column.data_type is StructuredColumnType.DATE:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            return date.fromisoformat(str(value).strip())
        if column.data_type is StructuredColumnType.DATETIME:
            if isinstance(value, datetime):
                parsed_datetime = value
            elif isinstance(value, date):
                parsed_datetime = datetime.combine(value, datetime.min.time())
            else:
                parsed_datetime = datetime.fromisoformat(str(value).strip())
            if parsed_datetime.tzinfo is not None and parsed_datetime.utcoffset() is not None:
                raise ValueError("timezone-aware datetimes are not supported")
            return parsed_datetime
        if column.data_type is StructuredColumnType.BOOLEAN:
            if isinstance(value, bool):
                return value
            normalized = str(value).strip().lower()
            if normalized in {"1", "true", "yes", "y", "on"}:
                return True
            if normalized in {"0", "false", "no", "n", "off"}:
                return False
            raise ValueError("expected a boolean")
    except (InvalidOperation, TypeError, ValueError, OverflowError) as error:
        raise StructuredIngestionError(
            "type_conversion_failed",
            f"Could not convert value to {column.data_type.value}.",
            row_number=row_number,
            physical_column=column.physical_name,
            sample=value,
        ) from error
    raise StructuredIngestionError(
        "type_conversion_failed",
        f"Unsupported confirmed type {column.data_type.value}.",
        row_number=row_number,
        physical_column=column.physical_name,
        sample=value,
    )


def _decimal_38_9(value: Any) -> Decimal:
    number = Decimal(str(value))
    if not number.is_finite() or (number and number.adjusted() > 28):
        raise ValueError("decimal value exceeds Decimal(38, 9)")
    with localcontext() as context:
        context.prec = 38
        quantized = number.quantize(Decimal("0.000000001"), rounding=ROUND_HALF_EVEN)
    if len(quantized.as_tuple().digits) > 38:
        raise ValueError("decimal value exceeds Decimal(38, 9)")
    return quantized


def _arrow_schema(columns: Sequence[StructuredColumnSchema]) -> pa.Schema:
    fields = [pa.field(column.physical_name, _arrow_type(column.data_type)) for column in columns]
    fields.extend(
        [
            pa.field("_source_id", pa.string(), nullable=False),
            pa.field("_dataset_id", pa.string(), nullable=False),
            pa.field("_schema_version", pa.uint64(), nullable=False),
            pa.field("_worksheet", pa.string(), nullable=False),
            pa.field("_row_number", pa.uint64(), nullable=False),
        ]
    )
    return pa.schema(fields)


def _arrow_type(column_type: StructuredColumnType) -> pa.DataType:
    return {
        StructuredColumnType.STRING: pa.string(),
        StructuredColumnType.INTEGER: pa.int64(),
        StructuredColumnType.DECIMAL: pa.decimal128(38, 9),
        StructuredColumnType.DATE: pa.date32(),
        StructuredColumnType.DATETIME: pa.timestamp("ms"),
        StructuredColumnType.BOOLEAN: pa.bool_(),
    }[column_type]


def _row_is_empty(values: Sequence[Any], formulas: Sequence[Any]) -> bool:
    return not any(value is not None and str(value).strip() for value in (*values, *formulas))


def _canonical_row(row: Mapping[str, Any]) -> bytes:
    return b"".join(_canonical_value(row[key]) for key in row)


def _canonical_value(value: Any) -> bytes:
    if value is None:
        return b"N;"
    if isinstance(value, datetime):
        rendered = value.isoformat(sep=" ", timespec="milliseconds")
    elif isinstance(value, date):
        rendered = value.isoformat()
    elif isinstance(value, Decimal):
        rendered = format(value, "f")
    elif isinstance(value, bool):
        rendered = "1" if value else "0"
    else:
        rendered = str(value)
    payload = rendered.encode("utf-8")
    return b"V" + str(len(payload)).encode("ascii") + b":" + payload + b";"


def _redacted_sample(value: Any) -> str:
    raw = str(value).encode("utf-8", errors="replace")
    digest = hashlib.sha256(raw).hexdigest()[:12]
    return f"[redacted sha256:{digest} len={len(raw)}]"


def _safe_path_component(value: str) -> str:
    normalized = _SAFE_PATH_COMPONENT.sub("_", value.strip())
    return normalized if normalized not in {"", ".", ".."} else "unknown"


def _require_within_root(root: Path, output_root: Path) -> None:
    if not output_root.resolve().is_relative_to(root.resolve()):
        raise StructuredIngestionError(
            "unsafe_parquet_path",
            "Parquet publication path must remain inside PARQUET_ROOT.",
        )


def _clear_publication_directory(output_root: Path, *, best_effort: bool = False) -> None:
    try:
        shutil.rmtree(output_root)
    except FileNotFoundError:
        return
    except OSError:
        if not best_effort:
            raise
