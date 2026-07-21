from __future__ import annotations

import codecs
import csv
import hashlib
import json
import re
from collections.abc import Iterable, Sequence
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_datetime as classify_datetime_format
from openpyxl.utils.exceptions import InvalidFileException

from .structured_models import (
    SpreadsheetPreview,
    StructuredColumnPreview,
    StructuredColumnType,
    StructuredDatasetPreview,
    StructuredDiagnostic,
)

MAX_SAMPLED_ROWS = 10_000
MAX_EXAMPLES = 5
MAX_WORKSHEETS = 64
MAX_COLUMNS_PER_DATASET = 256
MAX_LEADING_EMPTY_ROWS = 1_000
MAX_DIAGNOSTICS = 128
ENCODING_PROBE_BYTES = 64 * 1024
MAX_CSV_RECORD_BYTES = 1024 * 1024
MAX_CSV_FIELD_CHARS = 256 * 1024
MAX_SCANNED_ROWS = 50_000
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")
EXPECTED_WORKBOOK_ERRORS = (
    BadZipFile,
    EOFError,
    InvalidFileException,
    KeyError,
    OSError,
    ParseError,
    ValueError,
)
_INTEGER_RE = re.compile(r"^[+-]?\d+$")
_DECIMAL_RE = re.compile(r"^[+-]?(?:(?:\d+(?:\.\d*)?)|(?:\.\d+))(?:[eE][+-]?\d+)?$")
_PHYSICAL_RE = re.compile(r"[^a-z0-9_]+")


class _CsvRecordLimitExceeded(Exception):
    pass


class _BoundedCsvLines:
    def __init__(self, handle: Any, encoding: str) -> None:
        self.handle = handle
        self.encoding = encoding
        self.record_bytes = 0
        self.in_quotes = False

    def __iter__(self) -> _BoundedCsvLines:
        return self

    def __next__(self) -> str:
        remaining = MAX_CSV_RECORD_BYTES - self.record_bytes
        if remaining <= 0:
            raise _CsvRecordLimitExceeded
        line = self.handle.readline(remaining + 1)
        if not line:
            raise StopIteration
        line_bytes = len(line.encode(self.encoding))
        if line_bytes > remaining:
            raise _CsvRecordLimitExceeded
        self.record_bytes += line_bytes
        self.in_quotes = _updated_quote_state(line, self.in_quotes)
        if not self.in_quotes:
            self.record_bytes = 0
        return line


@dataclass(slots=True)
class _ColumnAccumulator:
    physical_name: str
    original_name: str
    display_name: str
    aliases: tuple[str, ...]
    observed_types: set[StructuredColumnType] = field(default_factory=set)
    examples: list[str] = field(default_factory=list)
    null_count: int = 0
    mixed_reported: bool = False

    def observe(
        self,
        value: Any,
        row_number: int,
        worksheet_name: str,
        diagnostics: list[StructuredDiagnostic],
    ) -> None:
        if _is_null(value):
            self.null_count += 1
            return

        value_type = _classify_value(value)
        self.observed_types.add(value_type)
        if not _types_are_compatible(self.observed_types) and not self.mixed_reported:
            self.mixed_reported = True
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="mixed_type",
                    message=(
                        f"Column {self.display_name!r} contains conflicting sampled values; "
                        "using string type."
                    ),
                    worksheet_name=worksheet_name,
                    column_name=self.display_name,
                    row_number=row_number,
                ),
            )
        rendered = _render_value(value)
        if rendered not in self.examples and len(self.examples) < MAX_EXAMPLES:
            self.examples.append(rendered)

    def preview(self, sampled_rows: int) -> StructuredColumnPreview:
        return StructuredColumnPreview(
            physical_name=self.physical_name,
            original_name=self.original_name,
            display_name=self.display_name,
            data_type=_resolved_type(self.observed_types),
            aliases=self.aliases,
            examples=tuple(self.examples),
            sampled_rows=sampled_rows,
            null_count=self.null_count,
        )


def infer_spreadsheet_schema(path: Path, source_id: str) -> SpreadsheetPreview:
    """Infer bounded schemas for a supported XLSX or CSV file."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".csv"}:
        raise ValueError(f"Unsupported spreadsheet format: {path.suffix or '<none>'}")

    diagnostics: list[StructuredDiagnostic] = []
    datasets: list[StructuredDatasetPreview] = []
    if suffix == ".csv":
        worksheet_name = path.stem
        encodings = _candidate_csv_encodings(path)
        if not encodings:
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="unsupported_encoding",
                    message="CSV could not be decoded with a supported encoding.",
                    worksheet_name=worksheet_name,
                ),
            )
            return SpreadsheetPreview(
                source_id=source_id, datasets=(), diagnostics=tuple(diagnostics)
            )
        for encoding in encodings:
            attempt_diagnostics: list[StructuredDiagnostic] = []
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    bounded_lines = _BoundedCsvLines(handle, encoding)
                    with _bounded_csv_field_size():
                        rows = (
                            (row_number, tuple(row), ())
                            for row_number, row in enumerate(csv.reader(bounded_lines), 1)
                        )
                        dataset = _infer_dataset(
                            source_id, worksheet_name, rows, attempt_diagnostics
                        )
            except UnicodeDecodeError:
                continue
            except _CsvRecordLimitExceeded:
                _add_diagnostic(
                    diagnostics,
                    StructuredDiagnostic(
                        code="csv_record_limit_exceeded",
                        message=f"CSV record exceeded {MAX_CSV_RECORD_BYTES} bytes.",
                        worksheet_name=worksheet_name,
                    ),
                )
                return SpreadsheetPreview(
                    source_id=source_id, datasets=(), diagnostics=tuple(diagnostics)
                )
            except csv.Error as exc:
                code = (
                    "csv_record_limit_exceeded"
                    if "field larger than field limit" in str(exc).lower()
                    else "csv_read_error"
                )
                _add_diagnostic(
                    diagnostics,
                    StructuredDiagnostic(
                        code=code,
                        message=f"CSV could not be read: {exc}",
                        worksheet_name=worksheet_name,
                    ),
                )
                return SpreadsheetPreview(
                    source_id=source_id, datasets=(), diagnostics=tuple(diagnostics)
                )
            except OSError as exc:
                _add_diagnostic(
                    diagnostics,
                    StructuredDiagnostic(
                        code="csv_read_error",
                        message=f"CSV could not be read: {exc}",
                        worksheet_name=worksheet_name,
                    ),
                )
                return SpreadsheetPreview(
                    source_id=source_id, datasets=(), diagnostics=tuple(diagnostics)
                )
            _merge_diagnostics(diagnostics, attempt_diagnostics)
            if dataset is not None:
                datasets.append(dataset)
            return SpreadsheetPreview(
                source_id=source_id,
                datasets=tuple(datasets),
                diagnostics=tuple(diagnostics),
            )
        _add_diagnostic(
            diagnostics,
            StructuredDiagnostic(
                code="unsupported_encoding",
                message="CSV decoding failed for every supported encoding.",
                worksheet_name=worksheet_name,
            ),
        )
        return SpreadsheetPreview(source_id=source_id, datasets=(), diagnostics=tuple(diagnostics))

    cached_workbook = formula_workbook = None
    try:
        try:
            cached_workbook = load_workbook(path, read_only=True, data_only=True)
            formula_workbook = load_workbook(path, read_only=True, data_only=False)
        except EXPECTED_WORKBOOK_ERRORS as exc:
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="workbook_read_error",
                    message=f"Workbook could not be opened: {exc}",
                    worksheet_name=path.stem,
                ),
            )
            return SpreadsheetPreview(
                source_id=source_id, datasets=(), diagnostics=tuple(diagnostics)
            )

        cached_sheets = cached_workbook.worksheets
        formula_sheets = {
            sheet.title: sheet for sheet in formula_workbook.worksheets[:MAX_WORKSHEETS]
        }
        if len(cached_sheets) > MAX_WORKSHEETS:
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="worksheet_limit_exceeded",
                    message=f"Only the first {MAX_WORKSHEETS} worksheets were sampled.",
                    worksheet_name=path.stem,
                ),
            )
        for sheet in cached_sheets[:MAX_WORKSHEETS]:
            try:
                formula_sheet = formula_sheets.get(sheet.title)
                cached_rows = (_xlsx_cached_values(row) for row in sheet.iter_rows())
                formula_rows = (
                    (_xlsx_formula_values(row) for row in formula_sheet.iter_rows())
                    if formula_sheet
                    else iter(())
                )
                paired_rows = (
                    (row_number, tuple(cached or ()), tuple(formulas or ()))
                    for row_number, (cached, formulas) in enumerate(
                        zip_longest(cached_rows, formula_rows, fillvalue=()), 1
                    )
                )
                dataset = _infer_dataset(source_id, sheet.title, paired_rows, diagnostics)
            except EXPECTED_WORKBOOK_ERRORS as exc:
                _add_diagnostic(
                    diagnostics,
                    StructuredDiagnostic(
                        code="sheet_read_error",
                        message=f"Worksheet could not be sampled: {exc}",
                        worksheet_name=sheet.title,
                    ),
                )
                continue
            if dataset is not None:
                datasets.append(dataset)
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()
    return SpreadsheetPreview(
        source_id=source_id, datasets=tuple(datasets), diagnostics=tuple(diagnostics)
    )


def _infer_dataset(
    source_id: str,
    worksheet_name: str,
    rows: Iterable[tuple[int, tuple[Any, ...], tuple[Any, ...]]],
    diagnostics: list[StructuredDiagnostic],
) -> StructuredDatasetPreview | None:
    iterator = iter(rows)
    header_row_number: int | None = None
    header_values: tuple[Any, ...] = ()
    header_formulas: tuple[Any, ...] = ()
    leading_empty_rows = 0
    for row_number, values, formulas in iterator:
        if _row_has_value(values, formulas):
            header_row_number = row_number
            header_values = values
            header_formulas = formulas
            break
        leading_empty_rows += 1
        if leading_empty_rows > MAX_LEADING_EMPTY_ROWS:
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="leading_empty_rows_exceeded",
                    message=f"Header was not found within {MAX_LEADING_EMPTY_ROWS} empty rows.",
                    worksheet_name=worksheet_name,
                    row_number=row_number,
                ),
            )
            return None
    if header_row_number is None:
        _add_diagnostic(
            diagnostics,
            StructuredDiagnostic(
                code="empty_sheet",
                message="Worksheet has no non-empty header row.",
                worksheet_name=worksheet_name,
            ),
        )
        return None

    raw_width = max(len(header_values), len(header_formulas))
    columns_truncated = raw_width > MAX_COLUMNS_PER_DATASET
    if columns_truncated:
        _add_diagnostic(
            diagnostics,
            StructuredDiagnostic(
                code="column_limit_exceeded",
                message=f"Only the first {MAX_COLUMNS_PER_DATASET} columns were sampled.",
                worksheet_name=worksheet_name,
                row_number=header_row_number,
            ),
        )
    width = min(raw_width, MAX_COLUMNS_PER_DATASET)
    accumulators = _build_columns(
        header_values, header_formulas, width, worksheet_name, header_row_number, diagnostics
    )
    for index, accumulator in enumerate(accumulators):
        cached_value = header_values[index] if index < len(header_values) else None
        formula = header_formulas[index] if index < len(header_formulas) else None
        if _is_formula(formula) and _is_null(cached_value):
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="formula_cache_missing",
                    message="Formula has no cached result; it was not evaluated.",
                    worksheet_name=worksheet_name,
                    column_name=accumulator.display_name,
                    row_number=header_row_number,
                ),
            )
    sampled_rows = 0
    scanned_rows = 0
    while sampled_rows < MAX_SAMPLED_ROWS:
        if scanned_rows >= MAX_SCANNED_ROWS:
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="scanned_row_limit_exceeded",
                    message=f"Sampling stopped after scanning {MAX_SCANNED_ROWS} data rows.",
                    worksheet_name=worksheet_name,
                ),
            )
            break
        try:
            row_number, values, formulas = next(iterator)
        except StopIteration:
            break
        scanned_rows += 1
        if not _row_has_value(values, formulas):
            continue
        row_width = max(len(values), len(formulas))
        if row_width > MAX_COLUMNS_PER_DATASET and not columns_truncated:
            columns_truncated = True
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="column_limit_exceeded",
                    message=f"Only the first {MAX_COLUMNS_PER_DATASET} columns were sampled.",
                    worksheet_name=worksheet_name,
                    row_number=row_number,
                ),
            )
        target_width = min(row_width, MAX_COLUMNS_PER_DATASET)
        if target_width > len(accumulators):
            _append_missing_columns(
                accumulators,
                target_width - len(accumulators),
                worksheet_name,
                header_row_number,
                diagnostics,
                sampled_rows,
            )
        sampled_rows += 1
        for index, accumulator in enumerate(accumulators):
            value = values[index] if index < len(values) else None
            formula = formulas[index] if index < len(formulas) else None
            if _is_formula(formula) and _is_null(value):
                _add_diagnostic(
                    diagnostics,
                    StructuredDiagnostic(
                        code="formula_cache_missing",
                        message="Formula has no cached result; it was not evaluated.",
                        worksheet_name=worksheet_name,
                        column_name=accumulator.display_name,
                        row_number=row_number,
                    ),
                )
            accumulator.observe(value, row_number, worksheet_name, diagnostics)

    columns = tuple(accumulator.preview(sampled_rows) for accumulator in accumulators)
    schema_hash = _schema_hash(columns)
    dataset_id = "ds-" + hashlib.sha256(f"{source_id}\0{worksheet_name}".encode()).hexdigest()[:24]
    return StructuredDatasetPreview(
        dataset_id=dataset_id,
        source_id=source_id,
        worksheet_name=worksheet_name,
        columns=columns,
        sampled_rows=sampled_rows,
        schema_hash=schema_hash,
    )


def _build_columns(
    values: Sequence[Any],
    formulas: Sequence[Any],
    width: int,
    worksheet_name: str,
    row_number: int,
    diagnostics: list[StructuredDiagnostic],
) -> list[_ColumnAccumulator]:
    accumulators: list[_ColumnAccumulator] = []
    used_names: set[str] = set()
    for index in range(width):
        raw = values[index] if index < len(values) else None
        if _is_null(raw):
            formula = formulas[index] if index < len(formulas) else None
            raw = formula if formula is not None and not _is_formula(formula) else None
        original = "" if _is_null(raw) else str(raw).strip()
        base = _normalize_physical_name(original, index + 1)
        physical = base
        if physical in used_names:
            suffix = 2
            while f"{base}_{suffix}" in used_names:
                suffix += 1
            physical = f"{base}_{suffix}"
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="duplicate_header",
                    message=f"Duplicate header {original or base!r} received a stable suffix.",
                    worksheet_name=worksheet_name,
                    column_name=original or base,
                    row_number=row_number,
                ),
            )
        elif not original:
            _add_diagnostic(
                diagnostics,
                StructuredDiagnostic(
                    code="missing_header",
                    message=f"Blank header received generated name {physical!r}.",
                    worksheet_name=worksheet_name,
                    column_name=physical,
                    row_number=row_number,
                ),
            )
        used_names.add(physical)
        accumulators.append(
            _ColumnAccumulator(
                physical_name=physical,
                original_name=original,
                display_name=original or physical,
                aliases=(original,) if original else (),
            )
        )
    return accumulators


def _append_missing_columns(
    accumulators: list[_ColumnAccumulator],
    count: int,
    worksheet_name: str,
    row_number: int,
    diagnostics: list[StructuredDiagnostic],
    sampled_rows: int,
) -> None:
    used_names = {column.physical_name for column in accumulators}
    start = len(accumulators) + 1
    for offset in range(count):
        index = start + offset
        base = _normalize_physical_name("", index)
        physical = base
        suffix = 2
        while physical in used_names:
            physical = f"{base}_{suffix}"
            suffix += 1
        used_names.add(physical)
        _add_diagnostic(
            diagnostics,
            StructuredDiagnostic(
                code="missing_header",
                message=f"Missing header received generated name {physical!r}.",
                worksheet_name=worksheet_name,
                column_name=physical,
                row_number=row_number,
            ),
        )
        accumulators.append(
            _ColumnAccumulator(
                physical_name=physical,
                original_name="",
                display_name=physical,
                aliases=(),
                null_count=sampled_rows,
            )
        )


def _detect_csv_encoding(path: Path) -> str | None:
    candidates = _candidate_csv_encodings(path)
    return candidates[0] if candidates else None


def _candidate_csv_encodings(path: Path) -> tuple[str, ...]:
    try:
        with path.open("rb") as handle:
            probe = handle.read(ENCODING_PROBE_BYTES)
    except OSError:
        return ()

    if probe.startswith((b"\xff\xfe", b"\xfe\xff", b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):
        return ()
    if b"\x00" in probe:
        return ()
    candidates: list[str] = []
    for encoding in CSV_ENCODINGS:
        decoder = codecs.getincrementaldecoder(encoding)(errors="strict")
        try:
            decoder.decode(probe, final=False)
        except (LookupError, UnicodeDecodeError):
            continue
        candidates.append(encoding)
    return tuple(candidates)


@contextmanager
def _bounded_csv_field_size():
    previous_limit = csv.field_size_limit()
    csv.field_size_limit(MAX_CSV_FIELD_CHARS)
    try:
        yield
    finally:
        csv.field_size_limit(previous_limit)


def _add_diagnostic(
    diagnostics: list[StructuredDiagnostic], diagnostic: StructuredDiagnostic
) -> None:
    if diagnostics and diagnostics[-1].code == "diagnostics_truncated":
        return
    if len(diagnostics) >= MAX_DIAGNOSTICS - 1:
        diagnostics.append(
            StructuredDiagnostic(
                code="diagnostics_truncated",
                message=f"Diagnostics were truncated at {MAX_DIAGNOSTICS} entries.",
                worksheet_name=diagnostic.worksheet_name,
            )
        )
        return
    diagnostics.append(diagnostic)


def _merge_diagnostics(
    target: list[StructuredDiagnostic], source: Sequence[StructuredDiagnostic]
) -> None:
    for diagnostic in source:
        _add_diagnostic(target, diagnostic)


def _normalize_physical_name(original: str, index: int) -> str:
    normalized = _PHYSICAL_RE.sub("_", original.lower()).strip("_")
    if normalized and normalized[0].isdigit():
        normalized = f"col_{normalized}"
    return normalized or f"column_{index}"


def _classify_value(value: Any) -> StructuredColumnType:
    if isinstance(value, bool):
        return StructuredColumnType.BOOLEAN
    if isinstance(value, datetime):
        return StructuredColumnType.DATETIME
    if isinstance(value, date):
        return StructuredColumnType.DATE
    if isinstance(value, int):
        return StructuredColumnType.INTEGER
    if isinstance(value, (float, Decimal)):
        return StructuredColumnType.DECIMAL
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"true", "false"}:
            return StructuredColumnType.BOOLEAN
        if _INTEGER_RE.fullmatch(text):
            return StructuredColumnType.INTEGER
        if _DECIMAL_RE.fullmatch(text):
            try:
                Decimal(text)
            except InvalidOperation:
                pass
            else:
                return StructuredColumnType.DECIMAL
        try:
            date.fromisoformat(text)
        except ValueError:
            pass
        else:
            return StructuredColumnType.DATE
        try:
            datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            pass
        else:
            return StructuredColumnType.DATETIME
    return StructuredColumnType.STRING


def _resolved_type(observed: set[StructuredColumnType]) -> StructuredColumnType:
    if not observed:
        return StructuredColumnType.STRING
    if observed == {StructuredColumnType.INTEGER}:
        return StructuredColumnType.INTEGER
    if observed <= {StructuredColumnType.INTEGER, StructuredColumnType.DECIMAL}:
        return StructuredColumnType.DECIMAL
    if len(observed) == 1:
        return next(iter(observed))
    return StructuredColumnType.STRING


def _types_are_compatible(observed: set[StructuredColumnType]) -> bool:
    return len(observed) <= 1 or observed <= {
        StructuredColumnType.INTEGER,
        StructuredColumnType.DECIMAL,
    }


def _render_value(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _schema_hash(columns: Sequence[StructuredColumnPreview]) -> str:
    normalized = [
        {
            "physical_name": column.physical_name,
            "original_name": column.original_name,
            "display_name": column.display_name,
            "data_type": column.data_type.value,
            "aliases": list(column.aliases),
        }
        for column in columns
    ]
    payload = json.dumps(normalized, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_null(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _row_has_value(values: Sequence[Any], formulas: Sequence[Any]) -> bool:
    return any(not _is_null(value) for value in values) or any(
        _is_formula(value) for value in formulas
    )


def _xlsx_cached_values(row: Sequence[Any]) -> tuple[Any, ...]:
    values: list[Any] = []
    for cell in row:
        value = cell.value
        if isinstance(value, datetime) and classify_datetime_format(cell.number_format) == "date":
            value = value.date()
        values.append(value)
    return _trim_trailing_empty(values)


def _xlsx_formula_values(row: Sequence[Any]) -> tuple[Any, ...]:
    return _trim_trailing_empty([cell.value for cell in row])


def _trim_trailing_empty(values: Sequence[Any]) -> tuple[Any, ...]:
    last_meaningful = 0
    for index, value in enumerate(values, 1):
        if not _is_null(value) or _is_formula(value):
            last_meaningful = index
    return tuple(values[:last_meaningful])


def _updated_quote_state(line: str, in_quotes: bool) -> bool:
    index = 0
    while index < len(line):
        if line[index] != '"':
            index += 1
            continue
        if in_quotes and index + 1 < len(line) and line[index + 1] == '"':
            index += 2
            continue
        in_quotes = not in_quotes
        index += 1
    return in_quotes
